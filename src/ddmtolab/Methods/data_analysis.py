"""
Data Analyzer Module for Multi-Task Optimization Experiments

This module provides a comprehensive analysis and visualization pipeline for
multi-task optimization experiments, including metric calculation, statistical
comparison tables, convergence plots, runtime analysis, and Pareto front visualization.

The statistical tests themselves are implemented in
:mod:`ddmtolab.Methods.statistical_tests`; this module reads the experiment
results, decides what to compare, and renders the outcome as tables and figures.

Classes:
    MetricResults: Dataclass for storing metric calculation results
    TableConfig: Dataclass for table generation configuration
    PlotConfig: Dataclass for plot generation configuration
    FriedmanResult: Dataclass for a Friedman test with post-hoc comparisons
    StatisticsCalculator: Statistics behind the tables and plots
    TableGenerator: Excel and LaTeX comparison tables
    PlotGenerator: Convergence, runtime, Pareto front and CD diagram figures
    DataAnalyzer: Main class for data analysis pipeline

Author: Jiangtao Shen
Email: j.shen5@exeter.ac.uk
Date: 2025.10.10
Version: 2.2
"""

import os
import pickle
import shutil
import warnings
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Union
from dataclasses import dataclass, field
from itertools import combinations
from enum import Enum

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from tqdm import tqdm

# Import from project modules
from ddmtolab.Methods.metrics import IGD, HV, GD, IGDp, FR, CV, DeltaP, Spread, Spacing
from ddmtolab.Methods.Algo_Methods.algo_utils import nd_sort
from ddmtolab.Methods import statistical_tests
from ddmtolab.Methods.statistical_tests import (  # noqa: F401  (re-exported)
    ALL_PAIRS_PROCEDURES,
    BERGMANN_MAX_ALGORITHMS,
    CLIFFS_DELTA_THRESHOLDS,
    EffectSizeResult,
    NemenyiComparison,
    NemenyiResult,
    OptimizationDirection,
    RankScheme,
)


# =============================================================================
# Enums and Constants
# =============================================================================

class TableFormat(Enum):
    """Output table format enumeration."""
    EXCEL = "excel"
    LATEX = "latex"


class StatisticType(Enum):
    """Statistical measure type enumeration."""
    MEAN = "mean"
    MEDIAN = "median"
    MAX = "max"
    MIN = "min"
    MEDIAN_IQR = "median_iqr"


# Default color palette for plots
DEFAULT_COLORS = [
    '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
    '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf',
    '#e41a1c', '#377eb8', '#4daf4a', '#984ea3', '#ff8c00'
]

# Default markers for plots
DEFAULT_MARKERS = ['o', 's', '^', 'v', 'D', 'p', '*', 'h', '<', '>', 'X', 'P', 'd', '8', 'H']

# =============================================================================
# Data Classes
# =============================================================================


@dataclass
class ScanResult:
    """
    Result of scanning a data directory.

    :no-index:

    Attributes
    ----------
    algorithms : List[str]
        Sorted list of algorithm names found in the directory.
    problems : List[str]
        Sorted list of problem names extracted from filenames.
    runs : int
        Number of independent runs per algorithm-problem combination.
    data_path : Path
        Path to the scanned data directory.
    """
    algorithms: List[str]
    problems: List[str]
    runs: int
    data_path: Path


@dataclass
class MetricResults:
    """
    Container for all metric calculation results.

    :no-index:

    Attributes
    ----------
    metric_values : Dict[str, Dict[str, Dict[int, List[np.ndarray]]]]
        Nested dictionary storing metric values per generation.
        Structure: metric_values[algorithm][problem][run] = List[np.ndarray]
        where each np.ndarray contains metric values per generation for each task.
    best_values : Dict[str, Dict[str, Dict[int, List[float]]]]
        Nested dictionary storing final best metric values.
        Structure: best_values[algorithm][problem][run] = List[float]
        where each float is the final best value for each task.
    objective_values : Dict[str, Dict[str, Dict[int, List[np.ndarray]]]]
        Nested dictionary storing original objective values.
        Structure: objective_values[algorithm][problem][run] = List[np.ndarray]
        where each np.ndarray has shape (n_solutions, n_objectives).
    runtime : Dict[str, Dict[str, Dict[int, float]]]
        Nested dictionary storing runtime in seconds.
        Structure: runtime[algorithm][problem][run] = float
    max_nfes : Dict[str, Dict[str, List[int]]]
        Nested dictionary storing maximum number of function evaluations.
        Structure: max_nfes[algorithm][problem] = List[int] (per task)
    metric_name : Optional[str]
        Name of the metric used (e.g., 'IGD', 'HV', or None for single-objective).
    """
    metric_values: Dict[str, Dict[str, Dict[int, Any]]]
    best_values: Dict[str, Dict[str, Dict[int, List[float]]]]
    objective_values: Dict[str, Dict[str, Dict[int, List[np.ndarray]]]]
    runtime: Dict[str, Dict[str, Dict[int, float]]]
    max_nfes: Dict[str, Dict[str, List[int]]]
    metric_name: Optional[str]


@dataclass
class TableConfig:
    """
    Configuration for table generation.

    :no-index:

    Attributes
    ----------
    table_format : TableFormat
        Output format (EXCEL or LATEX).
    statistic_type : StatisticType
        Type of statistic to display (MEAN, MEDIAN, MAX, MIN, MEDIAN_IQR).
    significance_level : float
        P-value threshold for statistical significance testing.
        Default: 0.05
    rank_sum_test : bool
        Whether to perform Wilcoxon rank-sum test.
        Default: True
    holm_correction : bool
        Whether to apply a Holm-Bonferroni correction to the per-instance
        rank-sum p-values. The family is every comparison against the baseline
        in one table generation, i.e. all non-baseline algorithms over all
        problem-task instances. When enabled the '+'/'-'/'=' symbols follow the
        corrected p-values instead of the raw ones.
        Default: False
    effect_size : bool
        Whether to report Cliff's delta alongside each comparison, shown in the
        cell as a separate bracketed field rather than folded into the symbol.
        Default: False
    friedman_test : bool
        Whether to append Friedman test rows (average ranks plus Holm-corrected
        post-hoc comparisons against a control algorithm) to the table.
        Default: False
    friedman_control : Optional[str]
        Control algorithm for the Friedman post-hoc comparisons. Defaults to the
        baseline, i.e. the last entry of the algorithm order.
        Default: None
    save_path : Path
        Directory path to save output tables.
    """
    table_format: TableFormat = TableFormat.EXCEL
    statistic_type: StatisticType = StatisticType.MEAN
    significance_level: float = 0.05
    rank_sum_test: bool = True
    holm_correction: bool = False
    effect_size: bool = False
    friedman_test: bool = False
    friedman_control: Optional[str] = None
    save_path: Path = Path('./Results')


@dataclass
class PlotConfig:
    """
    Configuration for plot generation.

    :no-index:

    Attributes
    ----------
    figure_format : str
        Output figure format (e.g., 'pdf', 'png', 'svg').
        Default: 'pdf'
    statistic_type : StatisticType
        Type of statistic for selecting representative run.
    log_scale : bool
        Whether to use logarithmic scale for y-axis.
        Default: False
    show_pf : bool
        Whether to show true Pareto front in ND solution plots.
        Default: True
    show_nd : bool
        Whether to filter and show only non-dominated solutions.
        Default: True
    merge_plots : bool
        Whether to merge all plots into a single figure.
        Default: False
    merge_columns : int
        Number of columns in merged plot layout.
        Default: 3
    show_std_band : bool
        Whether to show standard deviation band on convergence curves.
        Default: False
    save_path : Path
        Directory path to save output figures.
    colors : List[str]
        Color palette for plotting algorithms.
    markers : List[str]
        Marker styles for plotting algorithms.
    """
    figure_format: str = 'pdf'
    statistic_type: StatisticType = StatisticType.MEAN
    log_scale: bool = False
    show_pf: bool = True
    show_nd: bool = True
    merge_plots: bool = False
    merge_columns: int = 3
    show_std_band: bool = False
    save_path: Path = Path('./Results')
    colors: List[str] = field(default_factory=lambda: DEFAULT_COLORS.copy())
    markers: List[str] = field(default_factory=lambda: DEFAULT_MARKERS.copy())


@dataclass
class ComparisonResult:
    """
    Result of statistical comparison between algorithms.

    :no-index:

    Attributes
    ----------
    symbol : str
        Comparison symbol: '+' (better), '-' (worse), '=' (no significant difference).
    p_value : Optional[float]
        Raw (uncorrected) p-value from the statistical test, or None if the test
        was not performed.
    p_adjusted : Optional[float]
        Multiplicity-corrected p-value, populated only when a correction such as
        Holm-Bonferroni is applied over a family of comparisons. None otherwise.
    effect_size : Optional[float]
        Effect size (Cliff's delta) oriented so that a positive value means the
        algorithm is better than the baseline. None when not requested.
    effect_magnitude : Optional[str]
        Qualitative magnitude of ``effect_size``: 'negligible', 'small',
        'medium', 'large', or 'undefined'. None when not requested.
    """
    symbol: str
    p_value: Optional[float] = None
    p_adjusted: Optional[float] = None
    effect_size: Optional[float] = None
    effect_magnitude: Optional[str] = None


@dataclass
class FriedmanPostHocResult:
    """
    One post-hoc comparison of an algorithm against the control algorithm.

    :no-index:

    Attributes
    ----------
    algorithm : str
        Name of the compared algorithm.
    average_rank : float
        Average Friedman rank of the algorithm (1 is best).
    z_statistic : float
        Standardized rank difference against the control algorithm.
    p_value : Optional[float]
        Raw two-sided p-value, or None for the control algorithm itself.
    p_adjusted : Optional[float]
        Holm-Bonferroni corrected p-value over the k-1 comparisons, or None for
        the control algorithm itself.
    significant : bool
        Whether the corrected p-value is below the significance level.
    symbol : str
        '+' (better than control), '-' (worse), '=' (not significant), or ''
        for the control algorithm itself.
    """
    algorithm: str
    average_rank: float
    z_statistic: float
    p_value: Optional[float] = None
    p_adjusted: Optional[float] = None
    significant: bool = False
    symbol: str = ''


@dataclass
class FriedmanResult:
    """
    Outcome of a Friedman test with post-hoc comparisons against a control.

    :no-index:

    Attributes
    ----------
    statistic : float
        Friedman chi-squared statistic (chi^2_F).
    p_value : float
        P-value of the omnibus Friedman test.
    average_ranks : Dict[str, float]
        Average rank per algorithm (1 is best), keyed by algorithm name.
    n_algorithms : int
        Number of algorithms compared.
    n_instances : int
        Number of instances (blocks) retained after dropping incomplete ones.
    n_instances_dropped : int
        Number of instances dropped because at least one algorithm had no value.
    iman_davenport_statistic : float
        Iman-Davenport F_F statistic derived from chi^2_F, which corrects the
        Friedman statistic for being undesirably conservative.
    iman_davenport_p_value : float
        P-value of F_F under F(k-1, (k-1)(N-1)).
    control : Optional[str]
        Control algorithm used for the post-hoc comparisons.
    significance_level : float
        Significance level applied to the corrected post-hoc p-values.
    post_hoc : List[FriedmanPostHocResult]
        One entry per algorithm, in the order the algorithms were supplied.
    """
    statistic: float
    p_value: float
    average_ranks: Dict[str, float]
    n_algorithms: int
    n_instances: int
    n_instances_dropped: int = 0
    iman_davenport_statistic: float = np.nan
    iman_davenport_p_value: float = np.nan
    control: Optional[str] = None
    significance_level: float = 0.05
    post_hoc: List[FriedmanPostHocResult] = field(default_factory=list)


@dataclass
class ComparisonCounts:
    """
    Aggregated comparison counts for an algorithm.

    :no-index:

    Attributes
    ----------
    plus : int
        Number of significantly better results.
    minus : int
        Number of significantly worse results.
    equal : int
        Number of statistically equivalent results.
    """
    plus: int = 0
    minus: int = 0
    equal: int = 0


# =============================================================================
# Utility Functions
# =============================================================================

class DataUtils:
    """
    Utility class for data loading and processing operations.
    """

    @staticmethod
    def natural_sort_key(name: str) -> Tuple[int, int, str]:
        """
        Sort key ordering digit-bearing names numerically ('P2' before 'P10')
        while keeping digit-free names after them, alphabetically.

        Returns a homogeneous tuple so mixed name styles never raise the
        int-vs-str TypeError a naive conditional key produces.
        """
        digits = ''.join(filter(str.isdigit, name))
        if digits:
            return (0, int(digits), name)
        return (1, 0, name)

    @staticmethod
    def first_available_run(run_dict: Dict[int, Any]) -> Optional[int]:
        """
        Return the smallest run key present in a per-run dictionary, or None
        if the dictionary is empty (e.g. every run of a combination failed).

        Run numbering starts at 1 but run 1 may be missing when it failed, so
        callers must not hardcode it.
        """
        return min(run_dict.keys()) if run_dict else None

    @staticmethod
    def load_pickle(file_path: Path) -> Dict[str, Any]:
        """
        Load and return a Python object from a pickle file.

        Parameters
        ----------
        file_path : Path
            Path to the pickle file.

        Returns
        -------
        Dict[str, Any]
            Unpickled Python object (typically a dictionary containing
            'all_objs', 'runtime', 'max_nfes' keys).

        Raises
        ------
        FileNotFoundError
            If the pickle file does not exist.
        pickle.UnpicklingError
            If the file cannot be unpickled.
        """
        with open(file_path, 'rb') as f:
            return pickle.load(f)

    @staticmethod
    def load_reference(
            settings: Dict[str, Any],
            problem: str,
            task_identifier: Union[str, int],
            M: int,
            D: Optional[int] = None,
            C: int = 0
    ) -> Optional[np.ndarray]:
        """
        Load reference data (Pareto Front or reference point) for a specific problem and task.

        Parameters
        ----------
        settings : Dict[str, Any]
            Dictionary containing problem configurations and reference definitions.
            Expected keys:

            - problem (str): Contains task definitions
            - 'n_ref' (int, optional): Number of reference points (default: 10000)
            - 'ref_path' (str, optional): Path to reference files (default: './MOReference')

        problem : str
            Name of the problem (e.g., "DTLZ1", "DTLZ2").
        task_identifier : Union[str, int]
            Task identifier - either task name (str like "T1") or index (int like 0).
        M : int
            Number of objectives (required).
        D : int, optional
            Number of decision variables (dimension).
        C : int, optional
            Number of constraints (default: 0).

        Returns
        -------
        Optional[np.ndarray]
            Reference data with shape (n_points, M), or None if not available.

        Notes
        -----
        Supports three types of reference definitions:

        1. Callable: Function that returns reference data

           - Must accept parameter N (number of reference points)
           - Must accept parameter M (number of objectives)
           - May optionally accept parameters D, C
           - Example signatures: ``func(N, M)``, ``func(N, M, D)``, ``func(N, M, D, C)``

        2. String: File path to .npy or .csv reference file
        3. Array-like: Direct reference data (list, tuple, np.ndarray)

        If 'all_tasks' key is present instead of individual task keys, the same
        reference data will be used for all tasks.
        """
        # Convert task index to task name if necessary
        task_name = f"T{task_identifier + 1}" if isinstance(task_identifier, int) else task_identifier

        # Check if problem exists in settings
        if problem not in settings:
            warnings.warn(f"Problem '{problem}' not found in settings")
            return None

        problem_settings = settings[problem]

        # Check if task exists for this problem
        if task_name in problem_settings:
            ref_definition = problem_settings[task_name]
        elif 'all_tasks' in problem_settings:
            # Use the same reference for all tasks
            ref_definition = problem_settings['all_tasks']
        else:
            warnings.warn(f"Task '{task_name}' and 'all_tasks' not found for problem '{problem}'")
            return None

        # Case 1: Callable function
        if callable(ref_definition):
            N = settings.get('n_ref', 10000)

            try:
                import inspect
                sig = inspect.signature(ref_definition)
                params = list(sig.parameters.keys())
                num_params = len(params)

                if num_params == 2:
                    # func(N, M)
                    return ref_definition(N, M)
                elif num_params == 3:
                    # func(N, M, D)
                    if D is None:
                        warnings.warn(f"D not provided for {problem}_{task_name}, using 0")
                        D = 0
                    return ref_definition(N, M, D)
                elif num_params >= 4:
                    # func(N, M, D, C)
                    if D is None:
                        warnings.warn(f"D not provided for {problem}_{task_name}, using 0")
                        D = 0
                    return ref_definition(N, M, D, C)
                else:
                    warnings.warn(
                        f"Unexpected number of parameters ({num_params}) for "
                        f"reference function {problem}_{task_name}")
                    return None

            except Exception as e:
                warnings.warn(f"Failed to call reference function for {problem}_{task_name}: {e}")
                return None

        # Case 2: String (file path or file name)
        elif isinstance(ref_definition, str):
            return DataUtils._load_reference_from_file(
                settings,
                ref_definition,
                problem,
                task_name
            )

        # Case 3: Array-like (list, tuple, numpy array)
        elif isinstance(ref_definition, (list, tuple, np.ndarray)):
            reference = np.array(ref_definition)
            # Ensure it's at least 2D
            if reference.ndim == 1:
                reference = reference.reshape(1, -1)
            return reference

        else:
            warnings.warn(f"Unknown reference definition type for "
                          f"{problem}_{task_name}: {type(ref_definition)}")
            return None

    @staticmethod
    def _load_reference_from_file(
            settings: Dict[str, Any],
            ref_definition: str,
            problem: str,
            task_name: str
    ) -> Optional[np.ndarray]:
        """
        Load reference data from file.

        Parameters
        ----------
        settings : Dict[str, Any]
            Settings dictionary containing 'ref_path'.
        ref_definition : str
            File path or filename.
        problem : str
            Problem name for alternative path construction.
        task_name : str
            Task name for alternative path construction.

        Returns
        -------
        Optional[np.ndarray]
            Loaded reference data or None if loading fails.
        """
        ref_path = settings.get('ref_path', './MOReference')

        # Construct full path
        if not os.path.isabs(ref_definition):
            full_path = os.path.join(ref_path, ref_definition)
        else:
            full_path = ref_definition

        # Try to load the file
        try:
            if full_path.endswith('.npy'):
                return np.load(full_path)
            elif full_path.endswith('.csv'):
                return np.loadtxt(full_path, delimiter=',')
            else:
                print(f"Warning: Unsupported file format for '{full_path}'")
                return None
        except FileNotFoundError:
            # Try alternative naming conventions
            base_name = f"{problem}_{task_name}_ref"

            for ext in ['.npy', '.csv']:
                alt_path = os.path.join(ref_path, base_name + ext)
                if os.path.exists(alt_path):
                    try:
                        if ext == '.npy':
                            return np.load(alt_path)
                        else:
                            return np.loadtxt(alt_path, delimiter=',')
                    except Exception as e:
                        print(f"Error loading file '{alt_path}': {e}")

            print(f"Warning: File not found: '{full_path}'")
            return None
        except Exception as e:
            print(f"Error loading reference data from file '{full_path}': {e}")
            return None

    @staticmethod
    def get_metric_direction(metric_name: Optional[str]) -> OptimizationDirection:
        """
        Determine optimization direction based on metric type (Version 2 - More maintainable).

        Parameters
        ----------
        metric_name : Optional[str]
            Name of the metric or None for single-objective.

        Returns
        -------
        OptimizationDirection
            MINIMIZE or MAXIMIZE based on the metric's sign attribute.
        """
        if metric_name is None:
            return OptimizationDirection.MINIMIZE

        # Metric sign mapping (based on your code)
        # sign = -1 means minimize, sign = 1 means maximize
        metric_signs = {
            'IGD': -1,  # Inverted Generational Distance (minimize)
            'HV': 1,  # Hypervolume (maximize)
            'IGDp': -1,  # IGD+ (minimize)
            'GD': -1,  # Generational Distance (minimize)
            'DeltaP': -1,  # Delta_p (minimize)
            'Spacing': -1,  # Spacing (minimize)
            'Spread': -1,  # Spread (minimize)
            'FR': 1,  # Feasibility Rate (maximize)
            'CV': -1,  # Constraint Violation (minimize)
        }

        if metric_name not in metric_signs:
            raise ValueError(f'Unsupported metric: {metric_name}')

        sign = metric_signs[metric_name]
        return OptimizationDirection.MAXIMIZE if sign == 1 else OptimizationDirection.MINIMIZE


# =============================================================================
# Statistics Module
# =============================================================================

class StatisticsCalculator:
    """
    Statistics the results table and the plots are built from.

    The tests themselves live in :mod:`ddmtolab.Methods.statistical_tests`,
    which is the canonical entry point and the only place they are implemented.
    What remains here is what the reporting pipeline needs on top of them:

    - turning per-run samples into the displayed statistic
      (:meth:`calculate_statistic`) and into the ``+``/``-``/``=`` annotation
      (:meth:`perform_rank_sum_test`),
    - reshaping a test result into the dataclasses the table and the critical
      difference diagram render (:meth:`perform_friedman_test`,
      :meth:`perform_nemenyi_test`),
    - navigating the nested per-run result structure
      (:meth:`iterate_instances`, :meth:`build_instance_matrix`,
      :meth:`collect_task_data`, :meth:`select_representative_run`).

    The remaining ``perform_*`` and effect-size methods are thin wrappers kept
    for backward compatibility; new code should call the module functions
    directly, which follow the ``<name>_test`` naming and accept the same
    arguments in the same order for every test.
    """

    @staticmethod
    def calculate_statistic(
            data: List[float],
            statistic_type: StatisticType
    ) -> Tuple[float, Optional[float]]:
        """
        Calculate a statistical measure and optional standard deviation from data.

        Parameters
        ----------
        data : List[float]
            List of numeric values to compute statistics from.
        statistic_type : StatisticType
            Type of statistic to calculate (MEAN, MEDIAN, MAX, MIN, MEDIAN_IQR).

        Returns
        -------
        Tuple[float, Optional[float]]
            Tuple of (statistic_value, dispersion_value).
            The dispersion is the standard deviation for MEAN and the
            interquartile range for MEDIAN_IQR; it is None otherwise.
            Returns (np.nan, np.nan) for empty data.
        """
        if len(data) == 0:
            return np.nan, np.nan

        if statistic_type == StatisticType.MEAN:
            stat_value = np.mean(data)
            std_value = np.std(data, ddof=1) if len(data) > 1 else 0.0
            return stat_value, std_value
        elif statistic_type == StatisticType.MEDIAN:
            return np.median(data), None
        elif statistic_type == StatisticType.MEDIAN_IQR:
            q1, q3 = np.percentile(data, [25, 75])
            return np.median(data), q3 - q1
        elif statistic_type == StatisticType.MAX:
            return np.max(data), None
        elif statistic_type == StatisticType.MIN:
            return np.min(data), None
        else:
            return np.nan, np.nan

    @staticmethod
    def perform_rank_sum_test(
            algo_data: List[float],
            base_data: List[float],
            significance_level: float = 0.05,
            direction: OptimizationDirection = OptimizationDirection.MINIMIZE,
            compute_effect_size: bool = False
    ) -> ComparisonResult:
        """
        Compare the runs of two algorithms on one instance and render the
        verdict as a table symbol.

        The test itself is
        :func:`ddmtolab.Methods.statistical_tests.rank_sum_test`; this wrapper
        adds the ``+``/``-``/``=`` presentation the results table needs.

        Parameters
        ----------
        algo_data : List[float]
            Data from the algorithm being tested.
        base_data : List[float]
            Data from the baseline algorithm.
        significance_level : float, optional
            P-value threshold for significance (default: 0.05).
        direction : OptimizationDirection, optional
            Optimization direction (MINIMIZE or MAXIMIZE).
        compute_effect_size : bool, optional
            Whether to also compute Cliff's delta and store it in the
            ``effect_size`` / ``effect_magnitude`` fields of the result.
            Default: False, which leaves both fields as None.

        Returns
        -------
        ComparisonResult
            Result containing comparison symbol and p-value.
            Symbol: '+' (better), '-' (worse), '=' (no significant difference).
        """
        if len(algo_data) == 0 or len(base_data) == 0:
            return ComparisonResult(symbol='=', p_value=None)

        effect = None
        if compute_effect_size:
            effect = statistical_tests.cliffs_delta(algo_data, base_data, direction)

        try:
            p_value = statistical_tests.rank_sum_test(algo_data, base_data, direction).p_value

            if p_value < significance_level:
                algo_median = np.median(algo_data)
                base_median = np.median(base_data)

                if direction == OptimizationDirection.MINIMIZE:
                    symbol = '+' if algo_median < base_median else '-'
                else:
                    symbol = '+' if algo_median > base_median else '-'
            else:
                symbol = '='

            return ComparisonResult(
                symbol=symbol,
                p_value=p_value,
                effect_size=effect.delta if effect is not None else None,
                effect_magnitude=effect.magnitude if effect is not None else None
            )
        except Exception:
            return ComparisonResult(
                symbol='=',
                p_value=None,
                effect_size=effect.delta if effect is not None else None,
                effect_magnitude=effect.magnitude if effect is not None else None
            )

    @staticmethod
    def holm_bonferroni(p_values: List[Optional[float]]) -> List[Optional[float]]:
        """
        Apply the Holm-Bonferroni step-down correction to a family of p-values.

        The p-values are sorted ascending; the i-th smallest of m values is
        multiplied by ``m - i``, a running maximum enforces monotonicity, and the
        result is capped at 1. Entries that are None or NaN are not part of the
        family: they neither contribute to m nor receive an adjusted value.

        Because the correction is monotone non-decreasing, an adjusted p-value is
        never smaller than its raw counterpart, so a comparison can only lose
        significance through it, never gain it.

        Parameters
        ----------
        p_values : List[Optional[float]]
            Raw p-values of one family of comparisons, in any order.

        Returns
        -------
        List[Optional[float]]
            Adjusted p-values in the same order and of the same length as the
            input, with None wherever the input was None or NaN.

        Examples
        --------
        >>> StatisticsCalculator.holm_bonferroni([0.01, 0.02, 0.03])
        [0.03, 0.04, 0.04]
        """
        return statistical_tests.adjust_p_values(p_values, 'holm')

    @staticmethod
    def classify_cliffs_delta(delta: float) -> str:
        """
        Map a Cliff's delta to its qualitative magnitude.

        Thin wrapper over
        :func:`ddmtolab.Methods.statistical_tests.classify_cliffs_delta`.

        Parameters
        ----------
        delta : float
            Cliff's delta value; only its absolute value matters.

        Returns
        -------
        str
            'negligible' (``|delta| < 0.147``), 'small' (< 0.33), 'medium'
            (< 0.474), 'large' otherwise, or 'undefined' for NaN.
        """
        return statistical_tests.classify_cliffs_delta(delta)

    @staticmethod
    def cliffs_delta(
            algo_data: List[float],
            base_data: List[float],
            direction: OptimizationDirection = OptimizationDirection.MINIMIZE
    ) -> EffectSizeResult:
        """
        Compute Cliff's delta, a non-parametric effect size for two samples.

        Thin wrapper over
        :func:`ddmtolab.Methods.statistical_tests.cliffs_delta`, kept for the
        table pipeline and for backward compatibility.

        Parameters
        ----------
        algo_data : List[float]
            Data from the algorithm being tested. NaN entries are dropped.
        base_data : List[float]
            Data from the baseline algorithm. NaN entries are dropped.
        direction : OptimizationDirection, optional
            Optimization direction (MINIMIZE or MAXIMIZE).

        Returns
        -------
        EffectSizeResult
            Delta in [-1, 1] with its qualitative magnitude, or
            (np.nan, 'undefined') when either sample is empty.
        """
        return statistical_tests.cliffs_delta(algo_data, base_data, direction)

    @staticmethod
    def nemenyi_critical_value(n_algorithms: int, significance_level: float = 0.05) -> float:
        """
        Critical value q_alpha of the two-tailed Nemenyi test.

        Thin wrapper over
        :func:`ddmtolab.Methods.statistical_tests.nemenyi_critical_value`.

        Parameters
        ----------
        n_algorithms : int
            Number of algorithms compared (k >= 2).
        significance_level : float, optional
            Significance level alpha (default: 0.05).

        Returns
        -------
        float
            Critical value q_alpha.
        """
        return statistical_tests.nemenyi_critical_value(n_algorithms, significance_level)

    @staticmethod
    def perform_nemenyi_test(
            data_matrix: Union[np.ndarray, List[List[float]]],
            algorithm_names: List[str],
            direction: OptimizationDirection = OptimizationDirection.MINIMIZE,
            significance_level: float = 0.05
    ) -> NemenyiResult:
        """
        Run the Nemenyi all-pairs post-hoc test in critical-difference form.

        Thin wrapper over
        :func:`ddmtolab.Methods.statistical_tests.nemenyi_test`, kept because the
        critical difference diagram consumes its result.

        Parameters
        ----------
        data_matrix : Union[np.ndarray, List[List[float]]]
            Results with shape (n_algorithms, n_instances).
        algorithm_names : List[str]
            Algorithm names, one per row of ``data_matrix``.
        direction : OptimizationDirection, optional
            Optimization direction (MINIMIZE or MAXIMIZE).
        significance_level : float, optional
            Significance level of the critical difference (default: 0.05).

        Returns
        -------
        NemenyiResult
            Critical difference, average ranks, pairwise comparisons and the
            cliques the diagram connects.
        """
        return statistical_tests.nemenyi_test(
            data_matrix, algorithm_names, direction, significance_level
        )

    @staticmethod
    def perform_friedman_test(
            data_matrix: Union[np.ndarray, List[List[float]]],
            algorithm_names: List[str],
            direction: OptimizationDirection = OptimizationDirection.MINIMIZE,
            control: Optional[str] = None,
            significance_level: float = 0.05
    ) -> FriedmanResult:
        """
        Run a Friedman test with Holm-corrected post-hoc comparisons against a
        control algorithm, in the shape the results table renders.

        The statistics come from
        :func:`ddmtolab.Methods.statistical_tests.friedman_test` and the
        comparisons from
        :func:`ddmtolab.Methods.statistical_tests.control_post_hoc`; this wrapper
        only rearranges them into :class:`FriedmanResult` and adds the
        ``+``/``-``/``=`` symbols. For the aligned-ranks and Quade schemes, or
        for post-hoc procedures other than Holm, call those functions directly.

        Parameters
        ----------
        data_matrix : Union[np.ndarray, List[List[float]]]
            Results with shape (n_algorithms, n_instances); entry [i, j] is the
            performance of algorithm i on instance j.
        algorithm_names : List[str]
            Algorithm names, one per row of ``data_matrix``.
        direction : OptimizationDirection, optional
            Optimization direction (MINIMIZE or MAXIMIZE).
        control : Optional[str], optional
            Control algorithm for the post-hoc comparisons. Defaults to the last
            entry of ``algorithm_names``.
        significance_level : float, optional
            Threshold applied to the corrected post-hoc p-values (default: 0.05).

        Returns
        -------
        FriedmanResult
            Omnibus statistic and p-value, average ranks, and one post-hoc entry
            per algorithm, in the order the algorithms were supplied.

        Raises
        ------
        ValueError
            If fewer than 3 algorithms are supplied (the test is not applicable),
            if the number of names does not match the number of rows, if fewer
            than 2 complete instances remain, or if ``control`` is not one of
            ``algorithm_names``.
        """
        # The shape of the input is validated first, so a malformed call is
        # reported as such rather than as a too-small comparison
        ranking = statistical_tests.friedman_test(data_matrix, algorithm_names, direction)

        if len(algorithm_names) < 3:
            raise ValueError(
                f"The Friedman test needs at least 3 algorithms, got "
                f"{len(algorithm_names)}. Use perform_rank_sum_test for a "
                f"pairwise comparison instead."
            )

        control_name = control if control is not None else algorithm_names[-1]
        if control_name not in algorithm_names:
            raise ValueError(
                f"control '{control_name}' is not one of algorithm_names: {algorithm_names}"
            )

        average_ranks = ranking.average_ranks
        control_rank = average_ranks[control_name]

        family = statistical_tests.control_post_hoc(
            ranking, control_name, procedures=('holm',)
        )
        by_algorithm = {
            hypothesis.algorithms[0] if hypothesis.algorithms[1] == control_name
            else hypothesis.algorithms[1]: hypothesis
            for hypothesis in family.hypotheses
        }

        post_hoc = []
        for name in algorithm_names:
            if name == control_name:
                post_hoc.append(FriedmanPostHocResult(
                    algorithm=name,
                    average_rank=average_ranks[name],
                    z_statistic=0.0
                ))
                continue

            hypothesis = by_algorithm[name]
            p_adjusted = hypothesis.adjusted['holm']
            significant = p_adjusted is not None and p_adjusted < significance_level
            if significant:
                # A lower average rank is better, so it earns the '+'
                symbol = '+' if average_ranks[name] < control_rank else '-'
            else:
                symbol = '='

            post_hoc.append(FriedmanPostHocResult(
                algorithm=name,
                average_rank=average_ranks[name],
                z_statistic=hypothesis.z_statistic,
                p_value=hypothesis.p_value,
                p_adjusted=p_adjusted,
                significant=significant,
                symbol=symbol
            ))

        return FriedmanResult(
            statistic=ranking.statistic,
            p_value=ranking.p_value,
            average_ranks=average_ranks,
            n_algorithms=ranking.n_algorithms,
            n_instances=ranking.n_instances,
            n_instances_dropped=ranking.n_instances_dropped,
            iman_davenport_statistic=ranking.iman_davenport_statistic,
            iman_davenport_p_value=ranking.iman_davenport_p_value,
            control=control_name,
            significance_level=significance_level,
            post_hoc=post_hoc
        )

    @staticmethod
    def collect_task_data(
            all_best_values: Dict[str, Dict[str, Dict[int, List[float]]]],
            algo: str,
            prob: str,
            task_idx: int
    ) -> List[float]:
        """
        Collect non-NaN values from all runs for a specific algorithm-problem-task combination.

        Parameters
        ----------
        all_best_values : Dict[str, Dict[str, Dict[int, List[float]]]]
            Nested dictionary containing best metric values.
        algo : str
            Algorithm name.
        prob : str
            Problem name.
        task_idx : int
            Task index (0-based).

        Returns
        -------
        List[float]
            List of non-NaN metric values from all runs.
        """
        data = []
        for run in all_best_values[algo][prob].keys():
            value = all_best_values[algo][prob][run][task_idx]
            if not np.isnan(value):
                data.append(value)
        return data

    @staticmethod
    def iterate_instances(
            all_best_values: Dict[str, Dict[str, Dict[int, List[float]]]],
            algorithm_order: List[str]
    ) -> List[Tuple[str, int]]:
        """
        List every problem-task instance the algorithms are compared over.

        One instance is one row of the results table and one block of the
        Friedman and Nemenyi tests, so both go through this function and always
        see the same set. The task count of a problem comes from the first
        algorithm that has any run on it, so an algorithm whose runs all failed
        cannot make the problem disappear.

        Parameters
        ----------
        all_best_values : Dict[str, Dict[str, Dict[int, List[float]]]]
            Nested dictionary containing best metric values.
        algorithm_order : List[str]
            Algorithm names to consider.

        Returns
        -------
        List[Tuple[str, int]]
            Pairs of (problem name, 0-based task index), problems in natural
            order. Problems on which no algorithm has a single run are skipped
            with a warning.
        """
        problems = sorted(all_best_values[algorithm_order[0]].keys(),
                          key=DataUtils.natural_sort_key)

        instances = []
        for prob in problems:
            num_tasks = None
            for algo in algorithm_order:
                first_run = DataUtils.first_available_run(all_best_values[algo][prob])
                if first_run is not None:
                    num_tasks = len(all_best_values[algo][prob][first_run])
                    break

            if num_tasks is None:
                warnings.warn(f"No runs available for any algorithm on '{prob}', "
                              f"skipping it in the results table")
                continue

            instances.extend((prob, task_idx) for task_idx in range(num_tasks))

        return instances

    @staticmethod
    def build_instance_matrix(
            all_best_values: Dict[str, Dict[str, Dict[int, List[float]]]],
            algorithm_order: List[str],
            statistic_type: StatisticType = StatisticType.MEAN
    ) -> Tuple[np.ndarray, List[str]]:
        """
        Aggregate the runs of every algorithm into one value per instance.

        This is the matrix the rank-based omnibus tests consume: the runs are
        already collapsed into the same statistic the results table displays, so
        ranks derived from it agree with the numbers in the table.

        Parameters
        ----------
        all_best_values : Dict[str, Dict[str, Dict[int, List[float]]]]
            Nested dictionary containing best metric values.
        algorithm_order : List[str]
            Algorithm names, in the order the matrix rows should follow.
        statistic_type : StatisticType, optional
            Statistic used to collapse the runs of one instance
            (default: StatisticType.MEAN).

        Returns
        -------
        Tuple[np.ndarray, List[str]]
            Matrix of shape (n_algorithms, n_instances) and the instance labels,
            'P1' for a single-task problem and 'P1-T2' otherwise.
        """
        instances = StatisticsCalculator.iterate_instances(all_best_values, algorithm_order)

        multi_task = {prob for prob, task_idx in instances if task_idx > 0}
        labels = [f'{prob}-T{task_idx + 1}' if prob in multi_task else prob
                  for prob, task_idx in instances]

        matrix = np.array([
            [
                StatisticsCalculator.calculate_statistic(
                    StatisticsCalculator.collect_task_data(
                        all_best_values, algo, prob, task_idx
                    ),
                    statistic_type
                )[0]
                for prob, task_idx in instances
            ]
            for algo in algorithm_order
        ], dtype=float).reshape(len(algorithm_order), len(instances))

        return matrix, labels

    @staticmethod
    def select_representative_run(
            all_best_values: Dict[str, Dict[str, Dict[int, List[float]]]],
            algo: str,
            prob: str,
            task_idx: int,
            statistic_type: StatisticType
    ) -> Optional[int]:
        """
        Select a representative run based on the specified statistic type.

        Parameters
        ----------
        all_best_values : Dict[str, Dict[str, Dict[int, List[float]]]]
            Nested dictionary containing best metric values.
        algo : str
            Algorithm name.
        prob : str
            Problem name.
        task_idx : int
            Task index (0-based).
        statistic_type : StatisticType
            Type of statistic (MEAN returns None as all runs are used).
            MEDIAN_IQR selects the same run as MEDIAN.

        Returns
        -------
        Optional[int]
            Run number of the representative run, or None if MEAN or no valid data.
        """
        if statistic_type == StatisticType.MEAN:
            return None

        # Collect final values from all runs
        final_values = []
        runs = []

        for run in all_best_values[algo][prob].keys():
            value = all_best_values[algo][prob][run][task_idx]
            if not np.isnan(value):
                final_values.append(value)
                runs.append(run)

        if len(final_values) == 0:
            return None

        final_values = np.array(final_values)
        runs = np.array(runs)

        if statistic_type in (StatisticType.MEDIAN, StatisticType.MEDIAN_IQR):
            target_value = np.median(final_values)
            idx = np.argmin(np.abs(final_values - target_value))
        elif statistic_type == StatisticType.MAX:
            idx = np.argmax(final_values)
        elif statistic_type == StatisticType.MIN:
            idx = np.argmin(final_values)
        else:
            return None

        return runs[idx]


# =============================================================================
# Table Generator Module
# =============================================================================

class TableGenerator:
    """
    Class for generating comparison tables in Excel and LaTeX formats.
    """

    def __init__(self, config: TableConfig):
        """
        Initialize TableGenerator with configuration.

        Parameters
        ----------
        config : TableConfig
            Configuration object for table generation.
        """
        self.config = config

    def generate(
            self,
            all_best_values: Dict[str, Dict[str, Dict[int, List[float]]]],
            algorithm_order: List[str],
            metric_name: Optional[str] = None
    ) -> Union[pd.DataFrame, str]:
        """
        Generate comparison table with statistical analysis.

        Parameters
        ----------
        all_best_values : Dict[str, Dict[str, Dict[int, List[float]]]]
            Nested dictionary containing best metric values.
            Structure: all_best_values[algorithm][problem][run] = List[float]
        algorithm_order : List[str]
            List of algorithm names in display order.
            The last algorithm is treated as the baseline for comparisons.
        metric_name : Optional[str], optional
            Metric name to determine optimization direction.

        Returns
        -------
        Union[pd.DataFrame, str]
            DataFrame for Excel format, LaTeX string for LaTeX format.

        Raises
        ------
        ValueError
            If ``friedman_test`` is enabled but the test cannot be applied, for
            instance with fewer than 3 algorithms or fewer than 2 complete
            instances. The misconfiguration is surfaced rather than silently
            producing a table without the requested rows.
        """
        # Determine optimization direction
        direction = DataUtils.get_metric_direction(metric_name)

        # Generate data rows
        rows, comparison_counts, algorithm_ranks, instance_stats = self._generate_data_rows(
            all_best_values, algorithm_order, direction
        )

        friedman_result = None
        if self.config.friedman_test:
            friedman_result = self._run_friedman_test(instance_stats, algorithm_order, direction)

        # Generate and save table
        if self.config.table_format == TableFormat.EXCEL:
            return self._generate_excel_table(rows, algorithm_order, comparison_counts,
                                              algorithm_ranks, direction, friedman_result)
        else:
            return self._generate_latex_table(rows, algorithm_order, comparison_counts,
                                              algorithm_ranks, direction, friedman_result)

    def _run_friedman_test(
            self,
            instance_stats: List[Dict[str, float]],
            algorithm_order: List[str],
            direction: OptimizationDirection
    ) -> FriedmanResult:
        """
        Run the Friedman test over the instances of this table.

        The matrix handed to the test is the statistic already displayed in the
        cells, so the ranks in the table footer are consistent with the numbers
        above them.

        Parameters
        ----------
        instance_stats : List[Dict[str, float]]
            Per-instance statistic of every algorithm, one dict per table row.
        algorithm_order : List[str]
            Algorithm display order; its last entry is the default control.
        direction : OptimizationDirection
            Optimization direction (MINIMIZE or MAXIMIZE).

        Returns
        -------
        FriedmanResult
            Omnibus statistic, average ranks and post-hoc comparisons.

        Raises
        ------
        ValueError
            If the test cannot be applied to this table.
        """
        if not instance_stats:
            raise ValueError(
                "Friedman test requested but the table has no instances to rank."
            )

        matrix = np.array(
            [[row.get(algo, np.nan) for row in instance_stats] for algo in algorithm_order],
            dtype=float
        )
        return StatisticsCalculator.perform_friedman_test(
            matrix,
            list(algorithm_order),
            direction=direction,
            control=self.config.friedman_control,
            significance_level=self.config.significance_level
        )

    def _generate_data_rows(
            self,
            all_best_values: Dict[str, Dict[str, Dict[int, List[float]]]],
            algorithm_order: List[str],
            direction: OptimizationDirection
    ) -> Tuple[List[Dict[str, Any]], Dict[str, ComparisonCounts], Dict[str, List[int]],
               List[Dict[str, float]]]:
        """
        Build one table row per problem-task instance.

        The instances come from ``StatisticsCalculator.iterate_instances`` so the
        rows match the blocks of the omnibus tests exactly. Every instance is
        visited first and only formatted afterwards, because a Holm correction
        spans the whole family of comparisons in the table and can therefore not
        be resolved while the instances are still being read.

        Returns
        -------
        Tuple[List[Dict[str, Any]], Dict[str, ComparisonCounts], Dict[str, List[int]], List[Dict[str, float]]]
            Tuple of (rows, comparison_counts, algorithm_ranks, instance_stats).
            algorithm_ranks[algo] = List[int]; instance_stats holds the displayed
            statistic of every algorithm per instance, which feeds the Friedman test.
        """
        base_algo = algorithm_order[-1]
        instances = []
        comparison_counts = {algo: ComparisonCounts() for algo in algorithm_order[:-1]}
        algorithm_ranks = {algo: [] for algo in algorithm_order}

        for prob, task_idx in StatisticsCalculator.iterate_instances(
                all_best_values, algorithm_order):
            algo_stat_values = {}
            cells = {}

            base_data = StatisticsCalculator.collect_task_data(
                all_best_values, base_algo, prob, task_idx
            )

            for algo in algorithm_order:
                algo_data = StatisticsCalculator.collect_task_data(
                    all_best_values, algo, prob, task_idx
                )

                stat_value, std_value = StatisticsCalculator.calculate_statistic(
                    algo_data, self.config.statistic_type
                )

                algo_stat_values[algo] = stat_value

                result = None
                if self.config.rank_sum_test and algo != base_algo:
                    result = StatisticsCalculator.perform_rank_sum_test(
                        algo_data, base_data,
                        self.config.significance_level, direction,
                        compute_effect_size=self.config.effect_size
                    )

                cells[algo] = (stat_value, std_value, result)

            instances.append({
                'problem': prob,
                'task': task_idx + 1,
                'stats': algo_stat_values,
                'cells': cells
            })

        if self.config.holm_correction:
            self._apply_holm_correction(instances, algorithm_order)

        rows = []
        instance_stats = []

        for instance in instances:
            row = {'Problem': instance['problem'], 'Task': instance['task']}

            for algo in algorithm_order:
                stat_value, std_value, result = instance['cells'][algo]
                symbol = result.symbol if result is not None else ''

                if result is not None and algo in comparison_counts:
                    if symbol == '+':
                        comparison_counts[algo].plus += 1
                    elif symbol == '-':
                        comparison_counts[algo].minus += 1
                    else:
                        comparison_counts[algo].equal += 1

                row[algo] = self._format_cell_content(stat_value, std_value, symbol, result)

            row_ranks = self._calculate_row_ranks(instance['stats'], direction)
            for algo, rank in row_ranks.items():
                algorithm_ranks[algo].append(rank)

            rows.append(row)
            instance_stats.append(instance['stats'])

        return rows, comparison_counts, algorithm_ranks, instance_stats

    def _apply_holm_correction(
            self,
            instances: List[Dict[str, Any]],
            algorithm_order: List[str]
    ) -> None:
        """
        Holm-correct every comparison of this table in place.

        The family is all comparisons against the baseline over all instances of
        one table generation. Each affected ``ComparisonResult`` gets its
        ``p_adjusted`` filled in, and its symbol is relaxed to '=' when the
        corrected p-value no longer clears the significance level. The reverse
        can never happen, since a Holm-adjusted p-value is never smaller than the
        raw one, so the sign of a surviving symbol stays valid.

        Parameters
        ----------
        instances : List[Dict[str, Any]]
            Per-instance records produced by the first pass, modified in place.
        algorithm_order : List[str]
            Algorithm display order, used to visit the cells deterministically.

        Returns
        -------
        None
        """
        comparisons = [
            instance['cells'][algo][2]
            for instance in instances
            for algo in algorithm_order
            if instance['cells'][algo][2] is not None
        ]
        if not comparisons:
            return

        adjusted = StatisticsCalculator.holm_bonferroni(
            [comparison.p_value for comparison in comparisons]
        )

        for comparison, p_adjusted in zip(comparisons, adjusted):
            comparison.p_adjusted = p_adjusted
            if p_adjusted is None or p_adjusted >= self.config.significance_level:
                comparison.symbol = '='

    def _calculate_row_ranks(
            self,
            algo_values: Dict[str, float],
            direction: OptimizationDirection
    ) -> Dict[str, int]:
        """
        Calculate the rank of each algorithm in a single row.

        Parameters
        ----------
        algo_values : Dict[str, float]
            Statistical values for each algorithm.
        direction : OptimizationDirection
            Optimization direction (MINIMIZE or MAXIMIZE).

        Returns
        -------
        Dict[str, int]
            Rank of each algorithm (1 is the best).
        """
        # Filter out NaN values
        valid_algos = {algo: val for algo, val in algo_values.items() if not np.isnan(val)}

        if not valid_algos:
            return {algo: np.nan for algo in algo_values.keys()}

        # Sort based on optimization direction
        if direction == OptimizationDirection.MINIMIZE:
            sorted_algos = sorted(valid_algos.items(), key=lambda x: x[1])
        else:
            sorted_algos = sorted(valid_algos.items(), key=lambda x: x[1], reverse=True)

        # Assign ranks
        ranks = {}
        for rank, (algo, _) in enumerate(sorted_algos, start=1):
            ranks[algo] = rank

        # Set NaN for algorithms with NaN values
        for algo in algo_values.keys():
            if algo not in ranks:
                ranks[algo] = np.nan

        return ranks

    def _format_cell_content(
            self,
            stat_value: float,
            std_value: Optional[float],
            symbol: str,
            comparison: Optional[ComparisonResult] = None
    ) -> str:
        """
        Format a table cell with statistic value, optional dispersion, comparison
        symbol, and optional effect size.

        The dispersion is rendered as ``mean (std)`` for MEAN and ``median[IQR]``
        for MEDIAN_IQR. An effect size, when present, is appended as its own
        bracketed field so it is never read as part of the significance symbol.

        Parameters
        ----------
        stat_value : float
            Statistical value.
        std_value : Optional[float]
            Dispersion accompanying the statistic: standard deviation for MEAN,
            interquartile range for MEDIAN_IQR, None otherwise.
        symbol : str
            Comparison symbol.
        comparison : Optional[ComparisonResult], optional
            Comparison the cell belongs to, used only to render its effect size.
            Default: None, which appends nothing.

        Returns
        -------
        str
            Formatted cell content string.
        """
        if np.isnan(stat_value):
            return 'N/A'

        if self.config.table_format == TableFormat.EXCEL:
            if self.config.statistic_type == StatisticType.MEAN:
                cell_content = f"{stat_value:.3e} ({std_value:.1e})"
            elif self.config.statistic_type == StatisticType.MEDIAN_IQR:
                cell_content = f"{stat_value:.3e}[{std_value:.1e}]"
            else:
                cell_content = f"{stat_value:.3e}"

            if symbol:
                cell_content += f" {symbol}"
        else:
            # LaTeX format
            if self.config.statistic_type == StatisticType.MEAN:
                cell_content = f"${stat_value:.3e}$ (${std_value:.1e}$)"
            elif self.config.statistic_type == StatisticType.MEDIAN_IQR:
                cell_content = f"${stat_value:.3e}$[${std_value:.1e}$]"
            else:
                cell_content = f"${stat_value:.3e}$"

            if symbol:
                cell_content += f" ${symbol}$"

        cell_content += self._format_effect_size(comparison)

        return cell_content

    def _format_effect_size(self, comparison: Optional[ComparisonResult]) -> str:
        """
        Render the effect size of a comparison as a separate trailing field.

        Parameters
        ----------
        comparison : Optional[ComparisonResult]
            Comparison whose effect size should be rendered, if it carries one.

        Returns
        -------
        str
            Bracketed effect size prefixed by a space, or an empty string when
            no effect size was computed.
        """
        if comparison is None or comparison.effect_size is None:
            return ''

        delta = comparison.effect_size
        magnitude = comparison.effect_magnitude or 'undefined'

        if np.isnan(delta):
            return ' [d=N/A]'
        if self.config.table_format == TableFormat.EXCEL:
            return f" [d={delta:.2f} {magnitude}]"
        return f" [$\\delta={delta:.2f}$ {magnitude}]"

    def _find_best_value_in_row(
            self,
            row: Dict[str, Any],
            algorithm_order: List[str],
            direction: OptimizationDirection
    ) -> Optional[str]:
        """
        Find the algorithm with the best performance in a table row.

        Parameters
        ----------
        row : Dict[str, Any]
            Dictionary mapping algorithm names to formatted cell values.
        algorithm_order : List[str]
            List of algorithm names.
        direction : OptimizationDirection
            Optimization direction.

        Returns
        -------
        Optional[str]
            Name of the best-performing algorithm or None.
        """
        best_val = None
        best_algo = None

        for algo in algorithm_order:
            cell = row[algo]
            if cell != 'N/A':
                try:
                    # Cell formats: "1.2e+00", "1.2e+00 +", "1.2e+00 (3e-01) +",
                    # "1.2e+00[4e-01] +", a trailing " [d=0.42 large]" effect
                    # size, and LaTeX "$...$" variants -- strip dispersion,
                    # symbol and effect size parts
                    val_str = cell.split('(')[0].split('[')[0].replace('$', '').split()[0]
                    val = float(val_str)

                    if direction == OptimizationDirection.MINIMIZE:
                        if best_val is None or val < best_val:
                            best_val = val
                            best_algo = algo
                    else:
                        if best_val is None or val > best_val:
                            best_val = val
                            best_algo = algo
                except Exception:
                    pass

        return best_algo

    def _build_friedman_rows(
            self,
            friedman_result: FriedmanResult,
            algorithm_order: List[str]
    ) -> List[Dict[str, Any]]:
        """
        Render a Friedman result as table rows.

        Produces two rows: the average ranks, labelled with the omnibus
        chi-squared statistic, the Iman-Davenport correction of it and their
        p-values, and the Holm-corrected post-hoc p-values against the control
        algorithm.

        Parameters
        ----------
        friedman_result : FriedmanResult
            Result to render.
        algorithm_order : List[str]
            Algorithm display order.

        Returns
        -------
        List[Dict[str, Any]]
            Two row dictionaries keyed like the data rows.
        """
        is_latex = self.config.table_format == TableFormat.LATEX

        if is_latex:
            rank_label = (f"Friedman Rank ($\\chi^2_F={friedman_result.statistic:.4g}$, "
                          f"$p={friedman_result.p_value:.2e}$; "
                          f"$F_F={friedman_result.iman_davenport_statistic:.4g}$, "
                          f"$p={friedman_result.iman_davenport_p_value:.2e}$)")
        else:
            rank_label = (f"Friedman Rank (chi2={friedman_result.statistic:.4g}, "
                          f"p={friedman_result.p_value:.2e}; "
                          f"F_F={friedman_result.iman_davenport_statistic:.4g}, "
                          f"p={friedman_result.iman_davenport_p_value:.2e})")

        rank_row = {'Problem': rank_label, 'Task': ''}
        for algo in algorithm_order:
            rank = friedman_result.average_ranks[algo]
            rank_row[algo] = f"${rank:.2f}$" if is_latex else f"{rank:.2f}"

        post_hoc = {entry.algorithm: entry for entry in friedman_result.post_hoc}
        if is_latex:
            # '_' is a LaTeX control character, so escape it out of both the
            # label and the control algorithm's name
            control_label = str(friedman_result.control).replace('_', '-')
            p_label = f"Friedman $p_{{Holm}}$ (vs {control_label})"
        else:
            p_label = f"Friedman p_Holm (vs {friedman_result.control})"

        p_row = {'Problem': p_label, 'Task': ''}
        for algo in algorithm_order:
            entry = post_hoc.get(algo)
            if entry is None or entry.p_adjusted is None:
                p_row[algo] = 'Control'
            elif is_latex:
                p_row[algo] = f"${entry.p_adjusted:.2e}$ ${entry.symbol}$"
            else:
                p_row[algo] = f"{entry.p_adjusted:.2e} {entry.symbol}"

        return [rank_row, p_row]

    def _generate_excel_table(
            self,
            rows: List[Dict[str, Any]],
            algorithm_order: List[str],
            comparison_counts: Dict[str, ComparisonCounts],
            algorithm_ranks: Dict[str, List[int]],
            direction: OptimizationDirection,
            friedman_result: Optional[FriedmanResult] = None
    ) -> pd.DataFrame:
        """
        Generate and save a formatted Excel table.

        Parameters
        ----------
        rows : List[Dict[str, Any]]
            Table row data.
        algorithm_order : List[str]
            Algorithm display order.
        comparison_counts : Dict[str, ComparisonCounts]
            Comparison result counts.
        algorithm_ranks : Dict[str, List[int]]
            Per-instance ranks used for the Average Rank row.
        direction : OptimizationDirection
            Optimization direction.
        friedman_result : Optional[FriedmanResult], optional
            Friedman result to append as extra rows, or None to omit them.

        Returns
        -------
        pd.DataFrame
            DataFrame containing the table data.
        """
        num_summary_rows = 1  # Always has Average Rank row

        if self.config.rank_sum_test:
            summary_row = {'Problem': '+/-/=', 'Task': ''}
            for algo in algorithm_order[:-1]:
                counts = comparison_counts[algo]
                summary_row[algo] = f"{counts.plus}/{counts.minus}/{counts.equal}"
            summary_row[algorithm_order[-1]] = 'Base'
            rows.append(summary_row)
            num_summary_rows += 1

        # Average Rank must stay the last row, so Friedman goes just before it
        if friedman_result is not None:
            friedman_rows = self._build_friedman_rows(friedman_result, algorithm_order)
            rows.extend(friedman_rows)
            num_summary_rows += len(friedman_rows)

        avg_rank_row = {'Problem': 'Average Rank', 'Task': ''}
        for algo in algorithm_order:
            ranks = algorithm_ranks[algo]
            valid_ranks = [r for r in ranks if not np.isnan(r)]
            if valid_ranks:
                avg_rank = np.mean(valid_ranks)
                avg_rank_row[algo] = f"{avg_rank:.2f}"
            else:
                avg_rank_row[algo] = 'N/A'
        rows.append(avg_rank_row)

        # Create DataFrame
        df = pd.DataFrame(rows)
        columns = ['Problem', 'Task'] + algorithm_order
        df = df[columns]

        # Save and format
        save_dir = Path(self.config.save_path)
        save_dir.mkdir(parents=True, exist_ok=True)
        output_file = save_dir / f'results_table_{self.config.statistic_type.value}.xlsx'
        df.to_excel(output_file, index=False)

        # Apply Excel formatting
        self._apply_excel_formatting(output_file, df, algorithm_order, direction,
                                     num_summary_rows)

        print(f"Excel table saved to: {output_file}")
        return df

    def _apply_excel_formatting(
            self,
            output_file: Path,
            df: pd.DataFrame,
            algorithm_order: List[str],
            direction: OptimizationDirection,
            num_summary_rows: Optional[int] = None
    ) -> None:
        """
        Style the saved workbook and bold the best value of every data row.

        ``num_summary_rows`` tells the bolding logic how many trailing rows are
        footers rather than data; it is derived from the config when omitted.
        """
        wb = load_workbook(output_file)
        ws = wb.active

        # Define styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        normal_font = Font(name='Times New Roman', size=11)
        bold_font = Font(name='Times New Roman', size=11, bold=True)

        # Apply formatting and auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = normal_font

                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass

            ws.column_dimensions[column_letter].width = max_length + 2

        # Bold the best value in each data row
        if num_summary_rows is None:
            num_summary_rows = 1  # Always has Average Rank row
            if self.config.rank_sum_test:
                num_summary_rows += 1  # Add +/-/= row

        num_data_rows = len(df) - num_summary_rows

        for row_idx in range(2, num_data_rows + 2):
            best_val = None
            best_col = None

            for col_idx, algo in enumerate(algorithm_order, start=3):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value

                if cell_value and cell_value != 'N/A':
                    try:
                        # Handles "1.2e+00", "1.2e+00 +", "1.2e+00 (3e-01) +",
                        # "1.2e+00[4e-01] +" and a trailing " [d=0.42 large]"
                        val_str = str(cell_value).split('(')[0].split('[')[0].split()[0]
                        val = float(val_str)

                        if direction == OptimizationDirection.MINIMIZE:
                            if best_val is None or val < best_val:
                                best_val = val
                                best_col = col_idx
                        else:
                            if best_val is None or val > best_val:
                                best_val = val
                                best_col = col_idx
                    except Exception:
                        pass

            if best_col is not None:
                ws.cell(row=row_idx, column=best_col).font = bold_font

        # Bold the best (minimum) average rank
        avg_rank_row_idx = len(df) + 1  # Last row in the table
        best_avg_rank = None
        best_avg_rank_col = None

        for col_idx, algo in enumerate(algorithm_order, start=3):
            cell = ws.cell(row=avg_rank_row_idx, column=col_idx)
            cell_value = cell.value

            if cell_value and cell_value != 'N/A':
                try:
                    avg_rank = float(cell_value)
                    if best_avg_rank is None or avg_rank < best_avg_rank:
                        best_avg_rank = avg_rank
                        best_avg_rank_col = col_idx
                except Exception:
                    pass

        if best_avg_rank_col is not None:
            ws.cell(row=avg_rank_row_idx, column=best_avg_rank_col).font = bold_font

        wb.save(output_file)

    def _generate_latex_table(
            self,
            rows: List[Dict[str, Any]],
            algorithm_order: List[str],
            comparison_counts: Dict[str, ComparisonCounts],
            algorithm_ranks: Dict[str, List[int]],
            direction: OptimizationDirection,
            friedman_result: Optional[FriedmanResult] = None
    ) -> str:
        """
        Render the table as a standalone LaTeX document and save it.

        ``friedman_result`` adds the average-rank and post-hoc footer rows when
        supplied; the Average Rank row stays last, as in the Excel output.
        """
        df = pd.DataFrame(rows)

        # Build table structure
        num_cols = len(algorithm_order) + 2
        col_format = '|'.join(['c'] * num_cols)
        col_format = '|' + col_format + '|'

        # Initialize LaTeX document
        latex_str = "\\documentclass[]{article}\n"
        latex_str += "\\usepackage[margin=2cm]{geometry}\n"
        latex_str += "\\usepackage[table]{xcolor}\n"
        latex_str += "\\usepackage{graphicx}\n\n"
        latex_str += "\\newcommand{\\best}{\\cellcolor[rgb]{0.68,0.85,1.0}}\n\n"
        latex_str += "\\title{}\n"
        latex_str += "\\author{}\n\n"
        latex_str += "\\begin{document}\n"
        latex_str += "\\maketitle\n\n"
        latex_str += "\\begin{table*}[htbp]\n"
        latex_str += "\\renewcommand{\\arraystretch}{1.2}\n"
        latex_str += "\\centering\n"
        latex_str += "\\caption{Your caption here}\n"
        latex_str += "\\label{tab:results}\n"
        latex_str += "\\resizebox{1.0\\textwidth}{!}{\n"
        latex_str += f"\\begin{{tabular}}{{{col_format}}}\n"
        latex_str += "\\hline\n"

        # Header row
        header = "Problem & Task & " + " & ".join(algorithm_order) + " \\\\\n"
        latex_str += header
        latex_str += "\\hline\n"

        # Data rows
        for _, row in df.iterrows():
            best_algo = self._find_best_value_in_row(row, algorithm_order, direction)

            row_str = f"{row['Problem'].replace('_', '-')} & {row['Task']}"
            for algo in algorithm_order:
                cell = row[algo]
                if algo == best_algo:
                    cell = f"\\best {cell}"
                row_str += f" & {cell}"
            row_str += " \\\\\n"
            latex_str += row_str
            latex_str += "\\hline\n"

        # Summary row
        if self.config.rank_sum_test:
            summary_str = "\\multicolumn{2}{|c|}{$+$/$-$/$=$}"
            for algo in algorithm_order[:-1]:
                counts = comparison_counts[algo]
                summary_str += f" & ${counts.plus}$/${counts.minus}$/${counts.equal}$"
            summary_str += " & Base \\\\\n"
            latex_str += summary_str
            latex_str += "\\hline\n"

        # Friedman footer rows, kept above Average Rank like in the Excel output
        if friedman_result is not None:
            for friedman_row in self._build_friedman_rows(friedman_result, algorithm_order):
                friedman_str = "\\multicolumn{2}{|c|}{" + str(friedman_row['Problem']) + "}"
                for algo in algorithm_order:
                    friedman_str += f" & {friedman_row[algo]}"
                friedman_str += " \\\\\n"
                latex_str += friedman_str
                latex_str += "\\hline\n"

        # Average Rank row with best rank highlighted
        avg_rank_str = "\\multicolumn{2}{|c|}{Average Rank}"

        # Calculate average ranks and find the best
        avg_ranks = {}
        for algo in algorithm_order:
            ranks = algorithm_ranks[algo]
            valid_ranks = [r for r in ranks if not np.isnan(r)]
            if valid_ranks:
                avg_ranks[algo] = np.mean(valid_ranks)
            else:
                avg_ranks[algo] = np.nan

        # Find algorithm with best (minimum) average rank
        valid_avg_ranks = {algo: rank for algo, rank in avg_ranks.items() if not np.isnan(rank)}
        best_rank_algo = min(valid_avg_ranks, key=valid_avg_ranks.get) if valid_avg_ranks else None

        # Generate Average Rank row
        for algo in algorithm_order:
            if np.isnan(avg_ranks[algo]):
                cell_content = "N/A"
            else:
                cell_content = f"${avg_ranks[algo]:.2f}$"

            # Highlight the best rank with gray background
            if algo == best_rank_algo:
                cell_content = f"\\best {cell_content}"

            avg_rank_str += f" & {cell_content}"

        avg_rank_str += " \\\\\n"
        latex_str += avg_rank_str
        latex_str += "\\hline\n"

        latex_str += "\\end{tabular}}\n"
        latex_str += "\\end{table*}\n\n"
        latex_str += "\\end{document}\n"

        # Save to file
        save_dir = Path(self.config.save_path)
        save_dir.mkdir(parents=True, exist_ok=True)
        output_file = save_dir / f'results_table_{self.config.statistic_type.value}.tex'
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(latex_str)
        print(f"LaTeX table saved to: {output_file}")

        return latex_str


# =============================================================================
# Plot Generator Module
# =============================================================================

class PlotGenerator:
    """
    Class for generating various visualization plots.
    """

    def __init__(self, config: PlotConfig):
        """
        Initialize PlotGenerator with configuration.

        Parameters
        ----------
        config : PlotConfig
            Configuration object for plot generation.
        """
        self.config = config

    @staticmethod
    def _calculate_legend_fontsize(n_algorithms: int) -> int:
        """
        Calculate legend font size based on number of algorithms.

        Linear interpolation:
        - 2 algorithms → font size 14
        - 15 algorithms → font size 6

        Parameters
        ----------
        n_algorithms : int
            Number of algorithms.

        Returns
        -------
        int
            Calculated legend font size.
        """
        if n_algorithms <= 2:
            return 14
        elif n_algorithms >= 15:
            return 6
        else:
            # Linear interpolation: y = 14 - (14-6)/(15-2) * (x-2)
            return int(round(14 - (8 / 13) * (n_algorithms - 2)))

    def plot_convergence_curves(
            self,
            metric_values: Dict[str, Dict[str, Dict[int, Any]]],
            best_values: Dict[str, Dict[str, Dict[int, List[float]]]],
            max_nfes: Dict[str, Dict[str, List[int]]],
            algorithm_order: List[str],
            metric_name: Optional[str] = None
    ) -> None:
        """
        Generate and save convergence curve plots for all algorithms, problems, and tasks.

        Parameters
        ----------
        metric_values : Dict[str, Dict[str, Dict[int, Any]]]
            Metric values per generation.
            Structure: metric_values[algorithm][problem][run] = List[np.ndarray]
        best_values : Dict[str, Dict[str, Dict[int, List[float]]]]
            Best metric values for representative run selection.
        max_nfes : Dict[str, Dict[str, List[int]]]
            Maximum number of function evaluations per task.
            Structure: max_nfes[algorithm][problem] = List[int]
        algorithm_order : List[str]
            List of algorithm names to plot.
        metric_name : Optional[str], optional
            Metric name for y-axis label.

        Returns
        -------
        None
            Saves figures to disk.
        """
        problems = sorted(metric_values[algorithm_order[0]].keys())
        save_dir = Path(self.config.save_path)
        save_dir.mkdir(parents=True, exist_ok=True)

        if self.config.merge_plots:
            # Merged plot mode: all problems/tasks in one figure
            self._plot_merged_convergence(
                metric_values, best_values, max_nfes,
                algorithm_order, problems, metric_name, save_dir
            )
        else:
            # Separate plot mode: one figure per problem/task
            for prob in problems:
                first_run = DataUtils.first_available_run(best_values[algorithm_order[0]][prob])
                if first_run is None:
                    warnings.warn(f"No runs available for '{algorithm_order[0]}' on "
                                  f"'{prob}', skipping its convergence plots")
                    continue
                num_tasks = len(best_values[algorithm_order[0]][prob][first_run])

                for task_idx in range(num_tasks):
                    fig = self._create_convergence_figure(
                        num_tasks, metric_values, best_values, max_nfes,
                        algorithm_order, prob, task_idx, metric_name
                    )

                    if num_tasks == 1:
                        output_file = save_dir / f'{prob}.{self.config.figure_format}'
                    else:
                        output_file = save_dir / f'{prob}-Task{task_idx + 1}.{self.config.figure_format}'

                    fig.savefig(output_file, dpi=300, bbox_inches='tight')
                    plt.close(fig)

            print(f"All convergence plots saved to: {save_dir}")

    def _create_convergence_figure(
            self,
            num_tasks: int,
            metric_values: Dict,
            best_values: Dict,
            max_nfes: Dict,
            algorithm_order: List[str],
            prob: str,
            task_idx: int,
            metric_name: Optional[str],
            ax: Optional[plt.Axes] = None,
            show_legend: bool = True
    ) -> plt.Figure:
        """
        Create a single convergence curve figure.

        Parameters
        ----------
        num_tasks : int
            Total number of tasks.
        metric_values : Dict
            Metric values dictionary.
        best_values : Dict
            Best values dictionary.
        max_nfes : Dict
            Max NFEs dictionary.
        algorithm_order : List[str]
            Algorithm order.
        prob : str
            Problem name.
        task_idx : int
            Task index.
        metric_name : Optional[str]
            Metric name for label.
        ax : Optional[plt.Axes], optional
            Existing axes to plot on. If None, creates new figure.
        show_legend : bool, optional
            Whether to show legend. Default: True.

        Returns
        -------
        plt.Figure
            Matplotlib figure object (None if ax was provided).
        """
        fig = None
        if ax is None:
            fig, ax = plt.subplots(figsize=(5, 3.5))

        # Collect curve data for y-axis range and max NFEs for x-axis formatting
        all_curves = []
        actual_max_nfes = 0

        # Adaptive line width and marker size based on number of algorithms
        n_algos = len(algorithm_order)
        if n_algos <= 4:
            markersize, linewidth = 8, 2.5
        elif n_algos <= 6:
            markersize, linewidth = 7, 2.0
        else:
            markersize, linewidth = 6, 1.6

        for idx, algo in enumerate(algorithm_order):
            if self.config.show_std_band:
                mean_curve, std_curve = self._get_convergence_mean_std(
                    metric_values, algo, prob, task_idx
                )
                curve = mean_curve
            else:
                selected_run = StatisticsCalculator.select_representative_run(
                    best_values, algo, prob, task_idx, self.config.statistic_type
                )
                curve = self._get_convergence_curve(metric_values, algo, prob, task_idx, selected_run)

            if len(curve) == 0:
                continue

            all_curves.append(curve)

            nfes = max_nfes[algo][prob][task_idx]
            actual_max_nfes = max(actual_max_nfes, nfes)
            x = np.linspace(0, nfes, len(curve))
            marker_interval = max(1, len(curve) // 10)

            color = self.config.colors[idx % len(self.config.colors)]

            ax.plot(
                x, curve, label=algo,
                color=color,
                marker=self.config.markers[idx % len(self.config.markers)],
                markevery=marker_interval,
                markersize=markersize, linewidth=linewidth, linestyle='-', alpha=0.7
            )

            if self.config.show_std_band and len(std_curve) > 0:
                ax.fill_between(
                    x, curve - std_curve, curve + std_curve,
                    alpha=0.15, color=color
                )

        # Set axis labels
        y_label = metric_name if metric_name is not None else 'Objective Value'
        ax.set_xlabel('NFEs', fontsize=14)
        ax.set_ylabel(y_label, fontsize=14)

        title = f'{prob}' if num_tasks == 1 else f'{prob} - Task {task_idx + 1}'
        ax.set_title(title, fontsize=14)
        ax.tick_params(axis='both', which='major', labelsize=14)

        # Auto-adjust legend font size based on number of algorithms
        if show_legend:
            legend_fontsize = self._calculate_legend_fontsize(len(algorithm_order))
            ax.legend(loc='best', fontsize=legend_fontsize)

        ax.grid(True, alpha=0.2, linestyle='-')

        # Apply axis formatting after all settings are complete
        if self.config.log_scale:
            ax.set_yscale('log')
            # Check data range; use linear scale if range is too small
            if len(all_curves) > 0:
                all_data = np.concatenate([c for c in all_curves])
                y_min, y_max = np.nanmin(all_data), np.nanmax(all_data)

                # Log scale is invalid for non-positive values and ineffective
                # for less than one order of magnitude
                if y_min <= 0 or y_max / y_min < 10:
                    print(
                        f"Warning: Data range unsuitable for log scale "
                        f"({y_min:.4f} to {y_max:.4f}), using linear scale")
                    ax.set_yscale('linear')
                    self._apply_scientific_notation(ax, actual_xmax=actual_max_nfes)
                else:
                    # Use log scale, still need x-axis scientific notation
                    from matplotlib.ticker import LogFormatterSciNotation
                    ax.yaxis.set_major_formatter(LogFormatterSciNotation())
                    # Use scientific notation for x-axis if > 10000
                    if actual_max_nfes > 10000:
                        from matplotlib.ticker import ScalarFormatter
                        formatter = ScalarFormatter(useMathText=True)
                        formatter.set_scientific(True)
                        formatter.set_powerlimits((0, 0))
                        ax.xaxis.set_major_formatter(formatter)
                        ax.xaxis.major.formatter._useMathText = True
        else:
            # Apply scientific notation only for linear scale
            self._apply_scientific_notation(ax, actual_xmax=actual_max_nfes)

        # Disable minor ticks (must be called after set_yscale)
        ax.minorticks_off()

        if fig is not None:
            fig.tight_layout()
        return fig

    def _plot_merged_convergence(
            self,
            metric_values: Dict,
            best_values: Dict,
            max_nfes: Dict,
            algorithm_order: List[str],
            problems: List[str],
            metric_name: Optional[str],
            save_dir: Path
    ) -> None:
        """
        Create a merged figure with all convergence curves.

        Parameters
        ----------
        metric_values : Dict
            Metric values dictionary.
        best_values : Dict
            Best values dictionary.
        max_nfes : Dict
            Max NFEs dictionary.
        algorithm_order : List[str]
            Algorithm order.
        problems : List[str]
            List of problem names.
        metric_name : Optional[str]
            Metric name for label.
        save_dir : Path
            Directory to save the figure.
        """
        # Collect all subplot info (problem, task_idx)
        subplot_info = []
        for prob in problems:
            first_run = DataUtils.first_available_run(best_values[algorithm_order[0]][prob])
            if first_run is None:
                warnings.warn(f"No runs available for '{algorithm_order[0]}' on "
                              f"'{prob}', skipping it in the merged plot")
                continue
            num_tasks = len(best_values[algorithm_order[0]][prob][first_run])
            for task_idx in range(num_tasks):
                subplot_info.append((prob, task_idx, num_tasks))

        n_plots = len(subplot_info)
        if n_plots == 0:
            return

        n_cols = self.config.merge_columns
        n_rows = (n_plots + n_cols - 1) // n_cols

        # Create figure with subplots
        fig, axes = plt.subplots(n_rows, n_cols, figsize=(5 * n_cols, 3.5 * n_rows))

        # Flatten axes array for easy iteration
        if n_rows == 1 and n_cols == 1:
            axes = np.array([[axes]])
        elif n_rows == 1:
            axes = axes.reshape(1, -1)
        elif n_cols == 1:
            axes = axes.reshape(-1, 1)

        axes_flat = axes.flatten()

        # Plot each subplot
        for i, (prob, task_idx, num_tasks) in enumerate(subplot_info):
            ax = axes_flat[i]
            self._create_convergence_figure(
                num_tasks, metric_values, best_values, max_nfes,
                algorithm_order, prob, task_idx, metric_name,
                ax=ax, show_legend=False
            )

        # Hide unused subplots
        for i in range(n_plots, len(axes_flat)):
            axes_flat[i].set_visible(False)

        # Add single legend at the top of the figure
        # Collect handles from all subplots to ensure completeness, then sort by algorithm_order
        handle_dict = {}
        for ax_item in axes_flat[:n_plots]:
            for h, l in zip(*ax_item.get_legend_handles_labels()):
                if l not in handle_dict:
                    handle_dict[l] = h
        handles = [handle_dict[algo] for algo in algorithm_order if algo in handle_dict]
        labels = [algo for algo in algorithm_order if algo in handle_dict]
        legend_fontsize = 18

        # Calculate legend columns
        n_legend_cols = min(len(algorithm_order), 7)

        # First apply tight_layout to position subplots
        fig.tight_layout(h_pad=2.0, w_pad=1.5)

        # Get the top position of the first row of subplots (in figure coordinates)
        first_row_top = axes_flat[0].get_position().y1

        # Fixed padding between legend and first row (absolute size in cm)
        legend_padding_cm = 1.0
        fig_height_inch = fig.get_size_inches()[1]
        legend_padding = legend_padding_cm / 2.54 / fig_height_inch  # cm -> inch -> figure coords

        fig.legend(
            handles, labels,
            loc='lower center',
            ncol=n_legend_cols,
            fontsize=legend_fontsize,
            bbox_to_anchor=(0.5, first_row_top + legend_padding)
        )

        output_file = save_dir / f'convergence_merged.{self.config.figure_format}'
        fig.savefig(output_file, dpi=300, bbox_inches='tight')
        plt.close(fig)

        print(f"Merged convergence plot saved to: {output_file}")

    def _get_convergence_curve(
            self,
            metric_values: Dict,
            algo: str,
            prob: str,
            task_idx: int,
            run: Optional[int]
    ) -> np.ndarray:
        """
        Extract convergence curve for a specific configuration.

        Parameters
        ----------
        metric_values : Dict
            Metric values dictionary.
        algo : str
            Algorithm name.
        prob : str
            Problem name.
        task_idx : int
            Task index.
        run : Optional[int]
            Specific run number (None for mean across runs).

        Returns
        -------
        np.ndarray
            Convergence curve values.
        """
        if run is not None:
            return np.array(metric_values[algo][prob][run][task_idx])
        else:
            all_curves = []
            for r in metric_values[algo][prob].keys():
                curve = np.array(metric_values[algo][prob][r][task_idx])
                if len(curve) > 0:
                    all_curves.append(curve)

            if len(all_curves) == 0:
                return np.array([])

            min_len = min(len(c) for c in all_curves)
            truncated_curves = [c[:min_len] for c in all_curves]
            return np.mean(truncated_curves, axis=0)

    def _get_convergence_mean_std(
            self,
            metric_values: Dict,
            algo: str,
            prob: str,
            task_idx: int
    ) -> Tuple[np.ndarray, np.ndarray]:
        """
        Compute mean and standard deviation of convergence curves across all runs.

        Parameters
        ----------
        metric_values : Dict
            Metric values dictionary.
        algo : str
            Algorithm name.
        prob : str
            Problem name.
        task_idx : int
            Task index.

        Returns
        -------
        Tuple[np.ndarray, np.ndarray]
            (mean_curve, std_curve). Both empty arrays if no data.
        """
        all_curves = []
        for r in metric_values[algo][prob].keys():
            curve = np.array(metric_values[algo][prob][r][task_idx])
            if len(curve) > 0:
                all_curves.append(curve)

        if len(all_curves) < 2:
            return np.array([]), np.array([])

        min_len = min(len(c) for c in all_curves)
        truncated_curves = np.array([c[:min_len] for c in all_curves])
        return np.mean(truncated_curves, axis=0).ravel(), np.std(truncated_curves, axis=0, ddof=1).ravel()

    def _apply_scientific_notation(
            self,
            ax: plt.Axes,
            actual_xmax: Optional[float] = None,
            x_threshold: float = 10000,
            y_threshold: float = 1000
    ) -> None:
        """
        Apply scientific notation to axes if values exceed threshold.

        Parameters
        ----------
        ax : plt.Axes
            Matplotlib axes object.
        actual_xmax : Optional[float], optional
            Actual maximum x value from data (not affected by matplotlib padding).
            If None, uses ax.get_xlim()[1].
        x_threshold : float, optional
            Threshold for x-axis to use scientific notation. Default is 10000.
        y_threshold : float, optional
            Threshold for y-axis to use scientific notation. Default is 1000.
        """
        from matplotlib.ticker import ScalarFormatter

        # Use actual data max to avoid inconsistency from matplotlib padding
        xmax = actual_xmax if actual_xmax is not None else ax.get_xlim()[1]
        ymax = ax.get_ylim()[1]

        # X-axis: use scientific notation if > threshold
        if xmax > x_threshold:
            formatter = ScalarFormatter(useMathText=True)
            formatter.set_scientific(True)
            formatter.set_powerlimits((0, 0))
            ax.xaxis.set_major_formatter(formatter)
            ax.xaxis.major.formatter._useMathText = True

        # Y-axis: use scientific notation if > threshold
        if ymax > y_threshold:
            formatter = ScalarFormatter(useMathText=True)
            formatter.set_scientific(True)
            formatter.set_powerlimits((0, 0))
            ax.yaxis.set_major_formatter(formatter)
            ax.yaxis.major.formatter._useMathText = True

    def plot_runtime(
            self,
            runtime: Dict[str, Dict[str, Dict[int, float]]],
            algorithm_order: List[str]
    ) -> None:
        """
        Generate and save a bar plot showing average runtime comparison.

        Parameters
        ----------
        runtime : Dict[str, Dict[str, Dict[int, float]]]
            Runtime dictionary.
            Structure: runtime[algorithm][problem][run] = float (seconds)
        algorithm_order : List[str]
            List of algorithm names in display order.

        Returns
        -------
        None
            Saves figure to disk.
        """
        problems = sorted(runtime[algorithm_order[0]].keys(),
                          key=DataUtils.natural_sort_key)
        save_dir = Path(self.config.save_path)
        save_dir.mkdir(parents=True, exist_ok=True)

        fig, ax = plt.subplots(figsize=(6, 3.5))

        n_algorithms = len(algorithm_order)
        n_problems = len(problems)
        bar_width = 0.8 / n_algorithms
        x_groups = np.arange(n_problems)

        for idx, algo in enumerate(algorithm_order):
            means = []
            stds = []

            for prob in problems:
                runtimes = [runtime[algo][prob][run] for run in runtime[algo][prob].keys()]
                means.append(np.mean(runtimes))

                # Only calculate std if there are at least 2 data points
                if len(runtimes) > 1:
                    stds.append(np.std(runtimes, ddof=1))
                else:
                    stds.append(0.0)  # No error bar for single data point

            x_offset = x_groups + (idx - n_algorithms / 2 + 0.5) * bar_width

            ax.bar(
                x_offset, means, bar_width,
                yerr=stds, label=algo,
                color=self.config.colors[idx % len(self.config.colors)],
                alpha=0.8, capsize=2,
                error_kw={'linewidth': 1.2, 'ecolor': 'black', 'alpha': 0.6}
            )

        ax.set_ylabel('Runtime (s)', fontsize=12)
        ax.set_xticks(x_groups)
        ax.set_xticklabels(problems, fontsize=12)
        ax.tick_params(axis='both', which='major', labelsize=10)
        ax.legend(loc='best', fontsize=12, framealpha=0.7)
        ax.grid(True, axis='y', alpha=0.3, linestyle='-')

        fig.tight_layout()

        output_file = save_dir / f'runtime_comparison.{self.config.figure_format}'
        fig.savefig(output_file, dpi=300, bbox_inches='tight')
        plt.close(fig)

        print(f"Runtime plot saved to: {output_file}")

    def plot_cd_diagram(
            self,
            nemenyi_result: NemenyiResult,
            metric_name: Optional[str] = None,
            filename: str = 'cd_diagram'
    ) -> Path:
        """
        Draw the critical difference diagram of a Nemenyi test.

        The layout follows Figure 1(a) of Demsar (2006): the average ranks are
        plotted on a horizontal axis turned so that the best rank is on the
        right, the critical difference is shown as a bar above the axis, and
        groups of algorithms that are not significantly different are connected
        by a thick bar below it.

        Parameters
        ----------
        nemenyi_result : NemenyiResult
            Result of ``StatisticsCalculator.perform_nemenyi_test``.
        metric_name : Optional[str], optional
            Metric the ranks were computed on, used in the figure title.
            Default: None, which titles the figure without a metric.
        filename : str, optional
            Stem of the output file (default: 'cd_diagram').

        Returns
        -------
        Path
            Path of the saved figure.
        """
        average_ranks = nemenyi_result.average_ranks
        critical_difference = nemenyi_result.critical_difference

        # Display order of the columns decides the colours, rank order the layout
        display_order = list(average_ranks.keys())
        ordered = sorted(display_order, key=lambda name: average_ranks[name])

        # The axis spans the ranks, but never less than the 1..k it could reach
        low = min(1, int(np.floor(min(average_ranks.values()))))
        high = max(len(ordered), int(np.ceil(max(average_ranks.values()))))
        span = high - low

        # The better-ranked half goes right, the rest left; within each half the
        # outermost algorithm gets the shortest connector, so nothing crosses
        n_right = (len(ordered) + 1) // 2
        sides = [(ordered[:n_right], 'right'), (ordered[n_right:][::-1], 'left')]
        n_rows = max(n_right, len(ordered) - n_right)

        # The clique bars live between the axis and the first label row, so that
        # row is pushed down whenever a tall stack of them needs the space
        clique_spacing = 0.16
        clique_depth = 0.18 + clique_spacing * max(0, len(nemenyi_result.cliques) - 1)
        first_row = max(1.0, clique_depth + 0.35)

        fig, ax = plt.subplots(figsize=(6.5, max(2.0, 0.5 * (first_row + n_rows + 2))))
        ax.set_xlim(high + 0.02 * span, low - 0.02 * span)
        ax.set_ylim(-(first_row + n_rows - 0.3), 2.0)
        ax.set_axis_off()

        # Rank axis with a tick per integer rank
        ax.plot([low, high], [0, 0], color='black', linewidth=1.2)
        for tick in range(low, high + 1):
            ax.plot([tick, tick], [0, 0.12], color='black', linewidth=1.2)
            ax.text(tick, 0.2, str(tick), ha='center', va='bottom', fontsize=10)

        # Critical difference bar, anchored at the worst end of the axis. Very
        # few instances can make CD exceed the axis; the bar then spans it whole
        # and the label keeps carrying the true value
        bar_y = 1.15
        bar_end = max(low, high - critical_difference)
        ax.plot([high, bar_end], [bar_y, bar_y], color='black', linewidth=1.4)
        for end in (high, bar_end):
            ax.plot([end, end], [bar_y - 0.08, bar_y + 0.08],
                    color='black', linewidth=1.4)
        ax.text((high + bar_end) / 2, bar_y + 0.14,
                f'CD = {critical_difference:.3g}', ha='center', va='bottom', fontsize=10)

        # One connector per algorithm: down from its rank, then out to the label
        label_offset = 0.04 * span
        for names, side in sides:
            for row, name in enumerate(names):
                rank = average_ranks[name]
                depth = -(first_row + row)
                edge = low - label_offset if side == 'right' else high + label_offset
                color = self.config.colors[display_order.index(name) % len(self.config.colors)]

                ax.plot([rank, rank], [0, depth], color=color, linewidth=1.4)
                ax.plot([rank, edge], [depth, depth], color=color, linewidth=1.4)
                ax.text(
                    edge - 0.01 * span if side == 'right' else edge + 0.01 * span,
                    depth, f'{name} ({rank:.2f})',
                    ha='left' if side == 'right' else 'right',
                    va='center', fontsize=10, clip_on=False
                )

        # Groups that the test cannot tell apart. The bar overhangs its members
        # slightly so that a tight group stays visible as a bar
        overhang = 0.015 * span
        for index, clique in enumerate(nemenyi_result.cliques):
            clique_y = -0.18 - clique_spacing * index
            ax.plot(
                [average_ranks[clique[0]] - overhang, average_ranks[clique[-1]] + overhang],
                [clique_y, clique_y],
                color='black', linewidth=3.5, solid_capstyle='round'
            )

        metric_label = f' on {metric_name}' if metric_name else ''
        ax.set_title(
            f'Nemenyi test{metric_label} '
            f'($\\alpha$ = {nemenyi_result.significance_level:g}, '
            f'k = {nemenyi_result.n_algorithms}, '
            f'N = {nemenyi_result.n_instances})',
            fontsize=11
        )

        save_dir = Path(self.config.save_path)
        save_dir.mkdir(parents=True, exist_ok=True)
        output_file = save_dir / f'{filename}.{self.config.figure_format}'
        fig.savefig(output_file, dpi=300, bbox_inches='tight')
        plt.close(fig)

        print(f"Critical difference diagram saved to: {output_file}")
        return output_file

    def plot_nd_solutions(
            self,
            best_values: Dict[str, Dict[str, Dict[int, List[float]]]],
            objective_values: Dict[str, Dict[str, Dict[int, List[np.ndarray]]]],
            algorithm_order: List[str],
            settings: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Generate and save non-dominated solution plots.

        Parameters
        ----------
        best_values : Dict[str, Dict[str, Dict[int, List[float]]]]
            Best values for representative run selection.
        objective_values : Dict[str, Dict[str, Dict[int, List[np.ndarray]]]]
            Original objective values.
            Structure: objective_values[algorithm][problem][run] = List[np.ndarray]
            where each np.ndarray has shape (n_solutions, n_objectives).
        algorithm_order : List[str]
            List of algorithm names.
        settings : Optional[Dict[str, Any]], optional
            Problem settings for loading true Pareto fronts.

        Returns
        -------
        None
            Saves figures to disk.
        """
        nd_folder = Path(self.config.save_path) / 'ND_Solutions'
        nd_folder.mkdir(parents=True, exist_ok=True)

        problems = list(objective_values[algorithm_order[0]].keys())

        if self.config.merge_plots:
            # Merged plot mode: all algorithms for each problem/task in one figure
            self._plot_merged_nd_solutions(
                best_values, objective_values, algorithm_order, problems, settings, nd_folder
            )
        else:
            # Separate plot mode: one figure per algorithm/problem/task
            for algo in algorithm_order:
                for prob in problems:
                    first_run = DataUtils.first_available_run(objective_values[algo][prob])
                    if first_run is None:
                        warnings.warn(f"No runs available for '{algo}' on '{prob}', "
                                      f"skipping its ND plots")
                        continue
                    n_tasks = len(objective_values[algo][prob][first_run])

                    for task_idx in range(n_tasks):
                        first_run_objs = objective_values[algo][prob][first_run][task_idx]
                        n_objectives = first_run_objs.shape[1]

                        if n_objectives <= 1:
                            continue

                        selected_run = StatisticsCalculator.select_representative_run(
                            best_values, algo, prob, task_idx, self.config.statistic_type
                        )

                        if selected_run is None:
                            # MEAN mode (or no valid data): fall back to the
                            # first run that actually exists for this combo
                            selected_run = DataUtils.first_available_run(
                                objective_values[algo][prob])
                        if selected_run is None:
                            continue

                        objectives = objective_values[algo][prob][selected_run][task_idx]

                        if objectives.shape[0] == 0:
                            continue

                        # Filter non-dominated solutions if requested
                        if self.config.show_nd:
                            front_no, _ = nd_sort(objectives, objectives.shape[0])
                            nd_solutions = objectives[front_no == 1]
                        else:
                            nd_solutions = objectives

                        # Load true Pareto front if requested
                        true_pf = None
                        if self.config.show_pf and settings is not None:
                            true_pf = DataUtils.load_reference(settings, prob, task_idx, M=n_objectives)

                        # Create appropriate plot based on number of objectives
                        fig = self._create_nd_plot(
                            nd_solutions, true_pf, n_objectives, n_tasks,
                            prob, task_idx, algo)

                        # Save figure
                        if n_tasks == 1:
                            filename = f'{prob}-{algo}.{self.config.figure_format}'
                        else:
                            filename = f'{prob}-Task{task_idx + 1}-{algo}.{self.config.figure_format}'

                        fig.savefig(nd_folder / filename, dpi=300)
                        plt.close(fig)

            print(f"All non-dominated solutions plots saved to: {nd_folder}\n")

    def _plot_merged_nd_solutions(
            self,
            best_values: Dict,
            objective_values: Dict,
            algorithm_order: List[str],
            problems: List[str],
            settings: Optional[Dict[str, Any]],
            nd_folder: Path
    ) -> None:
        """
        Create merged figures for non-dominated solutions.

        Each merged figure contains all algorithms for a specific problem/task.

        Parameters
        ----------
        best_values : Dict
            Best values dictionary.
        objective_values : Dict
            Objective values dictionary.
        algorithm_order : List[str]
            Algorithm order.
        problems : List[str]
            List of problem names.
        settings : Optional[Dict[str, Any]]
            Problem settings.
        nd_folder : Path
            Output folder path.
        """
        for prob in problems:
            first_algo = algorithm_order[0]
            first_run = DataUtils.first_available_run(objective_values[first_algo][prob])
            if first_run is None:
                warnings.warn(f"No runs available for '{first_algo}' on '{prob}', "
                              f"skipping it in the merged ND plot")
                continue
            n_tasks = len(objective_values[first_algo][prob][first_run])

            for task_idx in range(n_tasks):
                # Get n_objectives for this specific task
                task_objs = objective_values[first_algo][prob][first_run][task_idx]
                n_objectives = task_objs.shape[1]

                if n_objectives <= 1:
                    continue

                n_algos = len(algorithm_order)
                # One column per algorithm, single row
                n_cols = n_algos
                n_rows = 1

                # Determine if 3D plot is needed
                is_3d = n_objectives == 3

                if is_3d:
                    fig = plt.figure(figsize=(4.5 * n_cols, 3.5 * n_rows))
                else:
                    fig, axes = plt.subplots(n_rows, n_cols, figsize=(4.5 * n_cols, 3.5 * n_rows))

                    # Flatten axes array
                    if n_cols == 1:
                        axes = np.array([axes])
                    axes_flat = axes.flatten()

                # Load true Pareto front once
                true_pf = None
                if self.config.show_pf and settings is not None:
                    true_pf = DataUtils.load_reference(settings, prob, task_idx, M=n_objectives)

                # Build subplot title prefix
                task_label = f'{prob} - Task {task_idx + 1}' if n_tasks > 1 else prob

                for idx, algo in enumerate(algorithm_order):
                    selected_run = StatisticsCalculator.select_representative_run(
                        best_values, algo, prob, task_idx, self.config.statistic_type
                    )
                    if selected_run is None:
                        selected_run = DataUtils.first_available_run(
                            objective_values[algo][prob])
                    if selected_run is None:
                        continue

                    objectives = objective_values[algo][prob][selected_run][task_idx]
                    if objectives.shape[0] == 0:
                        continue

                    # Filter non-dominated solutions
                    if self.config.show_nd:
                        front_no, _ = nd_sort(objectives, objectives.shape[0])
                        nd_solutions = objectives[front_no == 1]
                    else:
                        nd_solutions = objectives

                    # Subplot title: "P1 - Task 1 - RVEA" or "P1 - RVEA"
                    subplot_title = f'{task_label} - {algo}'

                    # Create subplot
                    if is_3d:
                        ax = fig.add_subplot(n_rows, n_cols, idx + 1, projection='3d')
                        self._plot_nd_subplot_3d(ax, nd_solutions, true_pf, subplot_title)
                    else:
                        ax = axes_flat[idx]
                        self._plot_nd_subplot_2d(ax, nd_solutions, true_pf, n_objectives, subplot_title)

                fig.tight_layout()

                # Save figure
                if n_tasks == 1:
                    filename = f'{prob}_merged.{self.config.figure_format}'
                else:
                    filename = f'{prob}-Task{task_idx + 1}_merged.{self.config.figure_format}'

                fig.savefig(nd_folder / filename, dpi=300, bbox_inches='tight')
                plt.close(fig)

        print(f"Merged non-dominated solutions plots saved to: {nd_folder}\n")

    def _plot_nd_subplot_2d(
            self,
            ax: plt.Axes,
            nd_solutions: np.ndarray,
            true_pf: Optional[np.ndarray],
            n_objectives: int,
            title: str
    ) -> None:
        """Plot 2D or parallel coordinates subplot for merged ND solutions."""
        if n_objectives == 2:
            if true_pf is not None and true_pf.shape[1] == 2:
                sort_idx = np.argsort(true_pf[:, 0])
                sorted_pf = true_pf[sort_idx]
                ax.scatter(sorted_pf[:, 0], sorted_pf[:, 1],
                           c='gray', s=2, linewidth=0.1, zorder=1)

            ax.scatter(nd_solutions[:, 0], nd_solutions[:, 1],
                       c='dodgerblue', s=60, alpha=0.8, edgecolors='black',
                       linewidth=0.8, zorder=2)

            ax.set_xlabel('$f_1$', fontsize=12)
            ax.set_ylabel('$f_2$', fontsize=12)
            ax.grid(True, alpha=0.2, linestyle='-')
        else:
            # Parallel coordinates
            for i in range(nd_solutions.shape[0]):
                ax.plot(range(n_objectives), nd_solutions[i, :],
                        'b-', alpha=0.3, linewidth=0.8)
            ax.set_xlabel('Objective', fontsize=12)
            ax.set_ylabel('Value', fontsize=12)
            ax.set_xticks(range(n_objectives))
            ax.set_xticklabels([rf'$f_{{{i + 1}}}$' for i in range(n_objectives)])
            ax.grid(True, alpha=0.3, linestyle='--')

        ax.set_title(title, fontsize=12)

    def _plot_nd_subplot_3d(
            self,
            ax: plt.Axes,
            nd_solutions: np.ndarray,
            true_pf: Optional[np.ndarray],
            title: str
    ) -> None:
        """Plot 3D subplot for merged ND solutions."""
        if true_pf is not None and true_pf.shape[1] == 3:
            ax.scatter(true_pf[:, 0], true_pf[:, 1], true_pf[:, 2],
                       c='gray', s=4, alpha=0.2, zorder=1, depthshade=True)

        ax.scatter(nd_solutions[:, 0], nd_solutions[:, 1], nd_solutions[:, 2],
                   c='dodgerblue', s=60, alpha=0.8, edgecolors='black',
                   linewidth=0.8, zorder=2, depthshade=True)

        ax.set_xlabel('$f_1$', fontsize=10)
        ax.set_ylabel('$f_2$', fontsize=10)
        ax.set_zlabel('$f_3$', fontsize=10)
        ax.set_title(title, fontsize=12)
        ax.view_init(elev=20, azim=60)

    def _create_nd_plot(
            self,
            nd_solutions: np.ndarray,
            true_pf: Optional[np.ndarray],
            n_objectives: int,
            n_tasks: int,
            prob: str,
            task_idx: int,
            algo: str
    ) -> plt.Figure:
        """
        Create a non-dominated solution plot.

        Parameters
        ----------
        nd_solutions : np.ndarray
            Non-dominated solutions array with shape (n_solutions, n_objectives).
        true_pf : Optional[np.ndarray]
            True Pareto front array.
        n_objectives : int
            Number of objectives.
        n_tasks : int
            Total number of tasks.
        prob : str
            Problem name.
        task_idx : int
            Task index.
        algo : str
            Algorithm name.

        Returns
        -------
        plt.Figure
            Matplotlib figure object.
        """
        fig = plt.figure(figsize=(4.5, 3.5))

        if n_objectives == 2:
            ax = fig.add_subplot(111)

            if true_pf is not None and true_pf.shape[1] == 2:
                sort_idx = np.argsort(true_pf[:, 0])
                sorted_pf = true_pf[sort_idx]
                ax.scatter(sorted_pf[:, 0], sorted_pf[:, 1],
                           c='gray', s=2, linewidth=0.1, label='True PF', zorder=1)

            ax.scatter(nd_solutions[:, 0], nd_solutions[:, 1],
                       c='dodgerblue', s=60, alpha=0.8, edgecolors='black',
                       linewidth=0.8, label='ND Solutions', zorder=2)

            ax.set_xlabel('$f_1$', fontsize=12)
            ax.set_ylabel('$f_2$', fontsize=12)
            ax.grid(True, alpha=0.2, linestyle='-')

        elif n_objectives == 3:
            ax = fig.add_subplot(111, projection='3d')

            if true_pf is not None and true_pf.shape[1] == 3:
                ax.scatter(true_pf[:, 0], true_pf[:, 1], true_pf[:, 2],
                           c='gray', s=4, alpha=0.2, label='True PF', zorder=1, depthshade=True)

            ax.scatter(nd_solutions[:, 0], nd_solutions[:, 1], nd_solutions[:, 2],
                       c='dodgerblue', s=60, alpha=0.8, edgecolors='black',
                       linewidth=0.8, label='ND Solutions', zorder=2, depthshade=True)

            ax.set_xlabel('$f_1$', fontsize=12)
            ax.set_ylabel('$f_2$', fontsize=12)
            ax.set_zlabel('$f_3$', fontsize=12)

            ax.view_init(elev=20, azim=60)

        else:
            # Parallel coordinates for many-objective
            ax = fig.add_subplot(111)

            for i in range(nd_solutions.shape[0]):
                ax.plot(range(n_objectives), nd_solutions[i, :],
                        'b-', alpha=0.3, linewidth=0.8)

            ax.set_xlabel('Objective', fontsize=12)
            ax.set_ylabel('Value', fontsize=12)
            ax.set_xticks(range(n_objectives))
            ax.set_xticklabels([rf'$f_{{{i + 1}}}$' for i in range(n_objectives)])
            ax.grid(True, alpha=0.3, linestyle='--')

        title = f'{prob} - {algo}' if n_tasks == 1 else f'{prob} - Task{task_idx + 1} - {algo}'
        plt.title(title, fontsize=10)
        plt.tight_layout()

        return fig


# =============================================================================
# Main Data Analyzer Class
# =============================================================================

class DataAnalyzer:
    """
    Main class for comprehensive data analysis and visualization of multi-task optimization experiments.

    This class provides a complete pipeline for:

    - Scanning data directories to detect algorithms, problems, and runs
    - Calculating performance metrics (IGD, HV, or objective values)
    - Generating statistical comparison tables (Excel or LaTeX)
    - Creating convergence curve plots
    - Visualizing runtime comparisons
    - Plotting non-dominated solutions

    Attributes
    ----------
    data_path : Path
        Path to the data directory containing experiment results.
    settings : Optional[Dict[str, Any]]
        Problem settings including reference definitions and metric configuration.
    algorithm_order : Optional[List[str]]
        Custom ordering of algorithms for display.
    table_config : TableConfig
        Configuration for table generation.
    plot_config : PlotConfig
        Configuration for plot generation.
    """

    def __init__(
            self,
            data_path: Union[str, Path] = './Data',
            settings: Optional[Dict[str, Any]] = None,
            algorithm_order: Optional[List[str]] = None,
            save_path: Union[str, Path] = './Results',
            table_format: str = 'excel',
            figure_format: str = 'pdf',
            statistic_type: str = 'mean',
            significance_level: float = 0.05,
            rank_sum_test: bool = True,
            holm_correction: bool = False,
            effect_size: bool = False,
            friedman_test: bool = False,
            friedman_control: Optional[str] = None,
            log_scale: bool = False,
            show_pf: bool = True,
            show_nd: bool = True,
            merge_plots: bool = False,
            merge_columns: int = 3,
            show_std_band: bool = False,
            best_so_far: bool = True,
            clear_results: bool = True,
            convergence_k: Optional[int] = None,
            cd_diagram: bool = False,
            cd_alpha: Optional[float] = None,
            multi_problem_report: bool = False,
            report_scheme: str = 'friedman',
            report_control: Optional[str] = None
    ):
        """
        Initialize DataAnalyzer with configuration parameters.

        Parameters
        ----------
        data_path : Union[str, Path], optional
            Path to data directory containing algorithm subdirectories.
            Each subdirectory should contain pickle files named: ALGO_problem_run.pkl
            Default: './Data'
        settings : Optional[Dict[str, Any]], optional
            Problem settings dictionary containing:

            - Problem names as keys (e.g., 'P1', 'P2')
            - Task definitions as nested dictionaries
            - 'metric': str ('IGD' or 'HV')
            - 'ref_path': str (path to reference files)
            - 'n_ref': int (number of reference points)

            Default: None (single-objective mode)
        algorithm_order : Optional[List[str]], optional
            Custom ordering of algorithms for display.
            The last algorithm is used as baseline for statistical tests.
            Default: None (alphabetical order)
        save_path : Union[str, Path], optional
            Directory path to save all output files.
            Default: './Results'
        table_format : str, optional
            Output table format: 'excel' or 'latex'.
            Default: 'excel'
        figure_format : str, optional
            Output figure format: 'pdf', 'png', 'svg', etc.
            Default: 'pdf'
        statistic_type : str, optional
            Type of statistic: 'mean', 'median', 'max', 'min', 'median_iqr'.
            'mean' renders as ``mean (std)`` and 'median_iqr' as ``median[IQR]``.
            Default: 'mean'
        significance_level : float, optional
            P-value threshold for statistical significance testing.
            Default: 0.05
        rank_sum_test : bool, optional
            Whether to perform Wilcoxon rank-sum test.
            Default: True
        holm_correction : bool, optional
            Whether to Holm-Bonferroni correct the rank-sum p-values over all
            comparisons against the baseline in the table. Symbols then follow
            the corrected p-values, which is stricter than the raw ones.
            Default: False
        effect_size : bool, optional
            Whether to report Cliff's delta next to each comparison, as its own
            bracketed field rather than folded into the '+'/'-'/'=' symbol.
            Default: False
        friedman_test : bool, optional
            Whether to append Friedman rows (average ranks plus Holm-corrected
            post-hoc comparisons against a control) to the table. Requires at
            least 3 algorithms and 2 instances, and raises otherwise.
            Default: False
        friedman_control : Optional[str], optional
            Control algorithm of the Friedman post-hoc comparisons. Defaults to
            the baseline, i.e. the last algorithm of the display order.
            Default: None
        log_scale : bool, optional
            Whether to use logarithmic scale for convergence plot y-axis.
            Default: False
        show_pf : bool, optional
            Whether to show true Pareto front in ND solution plots.
            Default: True
        show_nd : bool, optional
            Whether to filter and show only non-dominated solutions.
            Default: True
        merge_plots : bool, optional
            Whether to merge all plots into a single figure.
            Default: False
        merge_columns : int, optional
            Number of columns in merged plot layout.
            Default: 3
        show_std_band : bool, optional
            Whether to show standard deviation band on convergence curves.
            Default: False
        best_so_far : bool, optional
            Whether to use best-so-far metric values.
            Default: True
        clear_results : bool, optional
            Whether to clear existing results folder before analysis.
            Default: True
        convergence_k : Optional[int], optional
            Number of data points to sample from convergence curves for export.
            If None, no convergence data is exported.
            Default: None
        cd_diagram : bool, optional
            Whether to draw the critical difference diagram of a Nemenyi
            all-pairs post-hoc test, in the style of Demsar (2006, Figure 1a).
            Requires at least 3 algorithms and 2 instances.
            Default: False
        cd_alpha : Optional[float], optional
            Significance level of the critical difference. Demsar's own diagrams
            use 0.10, since the all-pairs test is conservative.
            Default: None, which reuses ``significance_level``.
        multi_problem_report : bool, optional
            Whether to run the full multi-problem analysis of Derrac et al.
            (2011) and save it as ``statistical_report.xlsx`` or ``.tex``:
            rankings under all three schemes, the omnibus tests, both post-hoc
            families with their adjusted p-values, contrast estimation and the
            plain pairwise tests.
            Default: False
        report_scheme : str, optional
            Ranking scheme the post-hoc sections of that report are derived
            from: 'friedman', 'aligned' or 'quade'.
            Default: 'friedman'
        report_control : Optional[str], optional
            Control algorithm of the 1xN post-hoc family.
            Default: None, which uses the best-ranked algorithm.
        """
        self.data_path = Path(data_path)
        self.settings = settings
        self.algorithm_order = algorithm_order
        self.best_so_far = best_so_far
        self.clear_results = clear_results

        # Parse enums
        stat_type = StatisticType(statistic_type)
        tbl_format = TableFormat(table_format)

        # Initialize configurations
        self.table_config = TableConfig(
            table_format=tbl_format,
            statistic_type=stat_type,
            significance_level=significance_level,
            rank_sum_test=rank_sum_test,
            holm_correction=holm_correction,
            effect_size=effect_size,
            friedman_test=friedman_test,
            friedman_control=friedman_control,
            save_path=Path(save_path)
        )

        self.plot_config = PlotConfig(
            figure_format=figure_format,
            statistic_type=stat_type,
            log_scale=log_scale,
            show_pf=show_pf,
            show_nd=show_nd,
            merge_plots=merge_plots,
            merge_columns=merge_columns,
            show_std_band=show_std_band,
            save_path=Path(save_path)
        )

        self.convergence_k = convergence_k
        self.cd_diagram = cd_diagram
        self.cd_alpha = cd_alpha if cd_alpha is not None else significance_level
        self.multi_problem_report = multi_problem_report
        self.report_scheme = RankScheme(report_scheme)
        self.report_control = report_control

        # Internal state
        self._scan_result: Optional[ScanResult] = None
        self._metric_results: Optional[MetricResults] = None

    def scan_data(self) -> ScanResult:
        """
        Scan the data directory to detect algorithms, problems, run counts.

        Returns
        -------
        ScanResult
            Dataclass containing:

            - algorithms: List[str] - Sorted list of algorithm names
            - problems: List[str] - Sorted list of problem names
            - runs: int - Number of independent runs
            - data_path: Path - Path to scanned directory

        Raises
        ------
        FileNotFoundError
            If data_path does not exist.
        ValueError
            If no algorithm directories or pickle files found.
        """
        algorithms = []
        problems = []
        runs_dict = {}

        for algo_dir in [d for d in self.data_path.iterdir() if d.is_dir()]:
            algo = algo_dir.name
            algorithms.append(algo)
            runs_dict[algo] = {}

            for pkl in algo_dir.glob('*.pkl'):
                stem = pkl.stem
                prefix = algo + '_'
                if stem.startswith(prefix):
                    remainder = stem[len(prefix):]
                    last_underscore = remainder.rfind('_')
                    if last_underscore > 0:
                        prob = remainder[:last_underscore]
                        runs_dict[algo].setdefault(prob, []).append(pkl)

                        if prob not in problems:
                            problems.append(prob)

        algorithms.sort()
        problems.sort(key=DataUtils.natural_sort_key)

        if not algorithms or not problems:
            raise ValueError(
                f"No experiment data found under '{self.data_path}'. Expected "
                f"one subdirectory per algorithm containing "
                f"'<algo>_<problem>_<run>.pkl' result files."
            )

        # Use the maximum run count across all combinations; missing runs are
        # skipped (with a warning) during metric calculation.
        run_counts = [len(v) for algo_runs in runs_dict.values() for v in algo_runs.values()]
        runs = max(run_counts) if run_counts else 0
        if run_counts and min(run_counts) != runs:
            warnings.warn(
                f"Uneven run counts across algorithm-problem combinations "
                f"(min={min(run_counts)}, max={runs}); missing runs will be skipped."
            )

        print(f"Found {len(algorithms)} algorithms: {algorithms}")
        print(f"Found {len(problems)} problems: {problems}")
        print(f"Run times: {runs}")

        self._scan_result = ScanResult(
            algorithms=algorithms,
            problems=problems,
            runs=runs,
            data_path=self.data_path
        )

        return self._scan_result

    def calculate_metrics(self) -> MetricResults:
        """
        Calculate metric values for all algorithms, problems, and runs.

        Returns
        -------
        MetricResults
            Dataclass containing all computed metrics:

            - metric_values: Metric values per generation
            - best_values: Final best metric values
            - objective_values: Original objective values
            - runtime: Runtime in seconds
            - max_nfes: Maximum function evaluations
            - metric_name: Name of metric used

        Raises
        ------
        RuntimeError
            If scan_data() has not been called.
        """
        if self._scan_result is None:
            self.scan_data()

        scan = self._scan_result
        algo_order = self.algorithm_order if self.algorithm_order else scan.algorithms
        unknown = [a for a in algo_order if a not in scan.algorithms]
        if unknown:
            raise ValueError(
                f"algorithm_order contains name(s) with no data on disk: {unknown}. "
                f"Algorithms found under '{self.data_path}': {scan.algorithms}"
            )
        metric_name = self.settings.get('metric') if self.settings else None

        # Initialize storage dictionaries
        all_values = {algo: {prob: {} for prob in scan.problems} for algo in algo_order}
        all_values_best_so_far = {algo: {prob: {} for prob in scan.problems} for algo in algo_order}
        all_best_values = {algo: {prob: {} for prob in scan.problems} for algo in algo_order}
        original_objective_values = {algo: {prob: {} for prob in scan.problems} for algo in algo_order}
        all_runtime = {algo: {prob: {} for prob in scan.problems} for algo in algo_order}
        all_max_nfes = {algo: {prob: None for prob in scan.problems} for algo in algo_order}

        total = len(algo_order) * len(scan.problems) * scan.runs
        pbar = tqdm(total=total, desc="Calculating metric values", dynamic_ncols=False, delay=0.2)

        for algo in algo_order:
            for prob in scan.problems:
                for run in range(1, scan.runs + 1):
                    pkl_file = f"{algo}_{prob}_{run}.pkl"
                    pkl_path = self.data_path / algo / pkl_file

                    # Failed/incomplete experiments produce no pkl file; skip
                    # them instead of aborting the whole analysis.
                    if not pkl_path.exists():
                        warnings.warn(f"Missing result file, skipping: {pkl_path}")
                        pbar.update(1)
                        continue

                    data = DataUtils.load_pickle(pkl_path)
                    metric_values, metric_values_best_bs = self._get_single_run_metric_value(data, prob)

                    all_values[algo][prob][run] = metric_values
                    all_values_best_so_far[algo][prob][run] = metric_values_best_bs

                    last_vals = [
                        np.asarray(task_arr).ravel()[-1] if len(task_arr) > 0 else np.nan
                        for task_arr in metric_values_best_bs
                    ]
                    all_best_values[algo][prob][run] = last_vals

                    last_objs = [
                        data['all_objs'][t][-1] if len(data['all_objs'][t]) > 0
                        else np.empty((0, 1))
                        for t in range(len(data['all_objs']))
                    ]
                    original_objective_values[algo][prob][run] = last_objs

                    all_runtime[algo][prob][run] = data['runtime']

                    if all_max_nfes[algo][prob] is None:
                        all_max_nfes[algo][prob] = data['max_nfes']

                    pbar.update(1)

        pbar.close()

        selected = all_values_best_so_far if self.best_so_far else all_values

        self._metric_results = MetricResults(
            metric_values=selected,
            best_values=all_best_values,
            objective_values=original_objective_values,
            runtime=all_runtime,
            max_nfes=all_max_nfes,
            metric_name=metric_name
        )

        return self._metric_results

    def _get_single_run_metric_value(
            self,
            data: Dict[str, Any],
            prob: str
    ) -> Tuple[List[np.ndarray], List[np.ndarray]]:
        """
        Calculate metric values for a single run.

        Parameters
        ----------
        data : Dict[str, Any]
            Loaded pickle data containing 'all_objs' key.
        prob : str
            Problem name for loading references.

        Returns
        -------
        Tuple[List[np.ndarray], List[np.ndarray]]
            Tuple of (metric_values, metric_values_best_so_far).
            Each is a list of arrays, one per task.
        """
        all_decs = data['all_decs']
        all_objs = data['all_objs']
        all_cons = data.get('all_cons', None)
        n_tasks = len(all_objs)
        n_gens_per_task = [len(all_objs[t]) for t in range(n_tasks)]

        metric_values = [np.zeros((n_gens_per_task[t], 1)) for t in range(n_tasks)]
        metric_values_best_so_far = [np.zeros((n_gens_per_task[t], 1)) for t in range(n_tasks)]

        for t in range(n_tasks):
            task_key = f'T{t + 1}'
            best_so_far = None

            reference = None

            if self.settings is not None and n_gens_per_task[t] > 0:

                M = all_objs[t][0].shape[1]
                D = all_decs[t][0].shape[1]
                C = all_cons[t][0].shape[1] if all_cons is not None else 0

                reference = DataUtils.load_reference(
                    self.settings,
                    prob,
                    task_key,
                    M=M,
                    D=D,
                    C=C
                )

            for gen in range(n_gens_per_task[t]):
                objs_tgen = all_objs[t][gen]
                cons_tgen = all_cons[t][gen] if all_cons is not None else None
                M = objs_tgen.shape[1]

                if M == 1:
                    metric_value = np.min(objs_tgen[:, 0])
                    sign = -1
                else:
                    if self.settings is None:
                        raise ValueError('Multi-objective metric calculation requires settings parameter')

                    metric_name = self.settings.get('metric')

                    if metric_name == 'IGD':
                        metric_instance = IGD()
                        metric_value = metric_instance.calculate(objs_tgen, reference)
                        sign = metric_instance.sign
                    elif metric_name == 'HV':
                        metric_instance = HV()
                        if reference is None:
                            metric_value = np.nan
                        # If reference is 1D or single row, treat as ref point; otherwise as PF
                        elif reference.ndim == 1 or reference.shape[0] == 1:
                            ref_point = reference.flatten()
                            metric_value = metric_instance.calculate(objs_tgen, reference=ref_point)
                        else:
                            metric_value = metric_instance.calculate(objs_tgen, pf=reference)
                        sign = metric_instance.sign
                    elif metric_name == 'IGDp':
                        metric_instance = IGDp()
                        metric_value = metric_instance.calculate(objs_tgen, reference)
                        sign = metric_instance.sign
                    elif metric_name == 'GD':
                        metric_instance = GD()
                        metric_value = metric_instance.calculate(objs_tgen, reference)
                        sign = metric_instance.sign
                    elif metric_name == 'DeltaP':
                        metric_instance = DeltaP()
                        metric_value = metric_instance.calculate(objs_tgen, reference)
                        sign = metric_instance.sign
                    elif metric_name == 'Spacing':
                        metric_instance = Spacing()
                        metric_value = metric_instance.calculate(objs_tgen)
                        sign = metric_instance.sign
                    elif metric_name == 'Spread':
                        metric_instance = Spread()
                        metric_value = metric_instance.calculate(objs_tgen, reference)
                        sign = metric_instance.sign
                    elif metric_name == 'FR':
                        if cons_tgen is None:
                            raise ValueError('FR metric requires constraint data, '
                                             'but all_cons is not available')
                        metric_instance = FR()
                        metric_value = metric_instance.calculate(cons_tgen)
                        sign = metric_instance.sign
                    elif metric_name == 'CV':
                        if cons_tgen is None:
                            raise ValueError('CV metric requires constraint data, '
                                             'but all_cons is not available')
                        metric_instance = CV()
                        metric_value = metric_instance.calculate(cons_tgen)
                        sign = metric_instance.sign
                    else:
                        raise ValueError(f'Unsupported metric: {metric_name}')

                metric_values[t][gen, 0] = metric_value

                # NaN metric values (e.g. HV without a reference) must not
                # poison the best-so-far curve for later generations
                if not np.isnan(metric_value):
                    if best_so_far is None:
                        best_so_far = metric_value
                    elif sign == -1:
                        best_so_far = min(best_so_far, metric_value)
                    else:
                        best_so_far = max(best_so_far, metric_value)

                metric_values_best_so_far[t][gen, 0] = best_so_far if best_so_far is not None else np.nan

        return metric_values, metric_values_best_so_far

    def generate_tables(self) -> Union[pd.DataFrame, str]:
        """
        Generate comparison tables with statistical analysis.

        Returns
        -------
        Union[pd.DataFrame, str]
            DataFrame for Excel format, LaTeX string for LaTeX format.

        Raises
        ------
        RuntimeError
            If calculate_metrics() has not been called.
        """
        if self._metric_results is None:
            self.calculate_metrics()

        algo_order = self.algorithm_order if self.algorithm_order else self._scan_result.algorithms

        table_gen = TableGenerator(self.table_config)
        return table_gen.generate(
            self._metric_results.best_values,
            algo_order,
            self._metric_results.metric_name
        )

    def generate_cd_diagram(self) -> NemenyiResult:
        """
        Run a Nemenyi all-pairs post-hoc test and draw its critical difference
        diagram, the visualization proposed by Demsar (2006, Figure 1a).

        The test ranks the algorithms on the same per-instance statistic the
        results table displays, so the diagram and the ``Average Rank`` row of
        the table tell the same story.

        Returns
        -------
        NemenyiResult
            Critical difference, average ranks, pairwise comparisons and the
            cliques the diagram connects.

        Raises
        ------
        ValueError
            If the test cannot be applied, for instance with fewer than 3
            algorithms or fewer than 2 instances.
        """
        if self._metric_results is None:
            self.calculate_metrics()

        algo_order = self.algorithm_order if self.algorithm_order else self._scan_result.algorithms
        metric_name = self._metric_results.metric_name

        matrix, _ = StatisticsCalculator.build_instance_matrix(
            self._metric_results.best_values,
            algo_order,
            self.table_config.statistic_type
        )

        result = StatisticsCalculator.perform_nemenyi_test(
            matrix,
            list(algo_order),
            direction=DataUtils.get_metric_direction(metric_name),
            significance_level=self.cd_alpha
        )

        PlotGenerator(self.plot_config).plot_cd_diagram(result, metric_name)
        return result

    def generate_statistical_report(self) -> Dict[str, Any]:
        """
        Run the full multi-problem analysis of Derrac et al. (2011) and save it.

        The report answers, in order, the questions that tutorial poses: how the
        algorithms rank overall, whether the ranking is significant at all,
        which algorithms differ from the control, which differ from each other,
        by how much they differ, and what a plain pairwise test says. It is
        written as ``statistical_report.xlsx`` (one sheet per section) or
        ``statistical_report.tex``, following ``table_format``.

        Returns
        -------
        Dict[str, Any]
            ``'rankings'`` maps each :class:`~ddmtolab.Methods.statistical_tests.RankScheme`
            to its ranking result, ``'control'`` and ``'all_pairs'`` hold the two
            post-hoc families, ``'contrast'`` the median-based estimators, and
            ``'tables'`` the rendered DataFrames.

        Raises
        ------
        ValueError
            If fewer than 3 algorithms are available, in which case the omnibus
            tests do not apply and the pairwise tests of
            :mod:`ddmtolab.Methods.statistical_tests` should be used directly.
        """
        if self._metric_results is None:
            self.calculate_metrics()

        algo_order = list(self.algorithm_order if self.algorithm_order
                          else self._scan_result.algorithms)
        metric_name = self._metric_results.metric_name
        direction = DataUtils.get_metric_direction(metric_name)

        matrix, _ = StatisticsCalculator.build_instance_matrix(
            self._metric_results.best_values, algo_order, self.table_config.statistic_type
        )

        rankings = {
            scheme: statistical_tests.omnibus_test(matrix, algo_order, direction, scheme)
            for scheme in RankScheme
        }
        ranking = rankings[self.report_scheme]

        control = statistical_tests.control_post_hoc(ranking, self.report_control)

        procedures = ALL_PAIRS_PROCEDURES
        if len(algo_order) > BERGMANN_MAX_ALGORITHMS:
            warnings.warn(
                f"Skipping the Bergmann-Hommel procedure for "
                f"{len(algo_order)} algorithms, which is beyond the "
                f"{BERGMANN_MAX_ALGORITHMS} its enumeration can handle."
            )
            procedures = tuple(p for p in procedures if p != 'bergmann')
        all_pairs = statistical_tests.all_pairs_post_hoc(ranking, procedures)

        contrast = statistical_tests.contrast_estimation(matrix, algo_order)

        tables = self._build_report_tables(matrix, algo_order, direction,
                                           rankings, control, all_pairs, contrast)
        self._write_statistical_report(tables)

        return {'rankings': rankings, 'control': control, 'all_pairs': all_pairs,
                'contrast': contrast, 'tables': tables}

    def _build_report_tables(
            self,
            matrix: np.ndarray,
            algorithm_order: List[str],
            direction: OptimizationDirection,
            rankings: Dict[RankScheme, Any],
            control: Any,
            all_pairs: Any,
            contrast: Any
    ) -> Dict[str, pd.DataFrame]:
        """
        Render every section of the statistical report as a DataFrame.

        Parameters
        ----------
        matrix : np.ndarray
            Algorithm-by-instance matrix the analysis ran on.
        algorithm_order : List[str]
            Algorithm display order.
        direction : OptimizationDirection
            Optimization direction, needed by the plain pairwise tests.
        rankings : Dict[RankScheme, RankingResult]
            Ranking result of each scheme.
        control : PostHocResult
            The 1xN family against the control algorithm.
        all_pairs : PostHocResult
            The NxN family.
        contrast : ContrastResult
            Median-based difference estimators.

        Returns
        -------
        Dict[str, pd.DataFrame]
            Section name to table, in the order they should be written.
        """
        rank_rows = []
        for algo in algorithm_order:
            rank_rows.append({
                'Algorithm': algo,
                **{scheme.value: rankings[scheme].average_ranks[algo] for scheme in RankScheme}
            })
        for label, attribute in (('Statistic', 'statistic'), ('p-value', 'p_value')):
            rank_rows.append({
                'Algorithm': label,
                **{scheme.value: getattr(rankings[scheme], attribute) for scheme in RankScheme}
            })
        friedman = rankings[RankScheme.FRIEDMAN]
        rank_rows.append({'Algorithm': 'Iman-Davenport F_F',
                          RankScheme.FRIEDMAN.value: friedman.iman_davenport_statistic})
        rank_rows.append({'Algorithm': 'Iman-Davenport p-value',
                          RankScheme.FRIEDMAN.value: friedman.iman_davenport_p_value})

        tables = {
            'Rankings': pd.DataFrame(rank_rows),
            f'Control ({control.control})': self._post_hoc_table(control),
            'All pairs': self._post_hoc_table(all_pairs),
            'Contrast estimation': pd.DataFrame(
                contrast.estimators, index=contrast.algorithms, columns=contrast.algorithms
            ).reset_index(names='Algorithm'),
            'Pairwise tests': self._pairwise_table(matrix, algorithm_order, direction),
        }
        return tables

    @staticmethod
    def _post_hoc_table(result: Any) -> pd.DataFrame:
        """
        Render a post-hoc family as one row per hypothesis.

        Parameters
        ----------
        result : PostHocResult
            Family to render.

        Returns
        -------
        pd.DataFrame
            Hypothesis, z statistic, unadjusted p-value and one column of
            adjusted p-values per procedure.
        """
        return pd.DataFrame([
            {
                'Hypothesis': hypothesis.label,
                'z': hypothesis.z_statistic,
                'Unadjusted p': hypothesis.p_value,
                **{procedure: hypothesis.adjusted[procedure]
                   for procedure in result.procedures}
            }
            for hypothesis in result.hypotheses
        ])

    @staticmethod
    def _pairwise_table(
            matrix: np.ndarray,
            algorithm_order: List[str],
            direction: OptimizationDirection
    ) -> pd.DataFrame:
        """
        Run the two plain pairwise tests on every pair of algorithms.

        These make no correction for multiplicity, so they answer "are these two
        different" rather than "which of the k algorithms differ"; the post-hoc
        sections above are what a multi-algorithm claim should rest on.

        Parameters
        ----------
        matrix : np.ndarray
            Algorithm-by-instance matrix.
        algorithm_order : List[str]
            Algorithm display order, matching the matrix rows.
        direction : OptimizationDirection
            Optimization direction.

        Returns
        -------
        pd.DataFrame
            One row per pair with the sign test and the Wilcoxon signed-rank
            test.
        """
        rows = []
        for first, second in combinations(range(len(algorithm_order)), 2):
            signs = statistical_tests.sign_test(matrix[first], matrix[second], direction)
            wilcoxon = statistical_tests.wilcoxon_signed_rank_test(
                matrix[first], matrix[second], direction)
            rows.append({
                'Comparison': f'{algorithm_order[first]} vs {algorithm_order[second]}',
                'Wins': signs.wins,
                'Losses': signs.losses,
                'Ties': signs.ties,
                'Sign p': signs.p_value,
                'R+': wilcoxon.r_plus,
                'R-': wilcoxon.r_minus,
                'Wilcoxon p': wilcoxon.p_value,
            })
        return pd.DataFrame(rows)

    def _write_statistical_report(self, tables: Dict[str, pd.DataFrame]) -> Path:
        """
        Write the report sections to one workbook or one LaTeX file.

        Parameters
        ----------
        tables : Dict[str, pd.DataFrame]
            Section name to table.

        Returns
        -------
        Path
            Path of the written file.
        """
        save_dir = Path(self.table_config.save_path)
        save_dir.mkdir(parents=True, exist_ok=True)

        if self.table_config.table_format == TableFormat.EXCEL:
            output_file = save_dir / 'statistical_report.xlsx'
            with pd.ExcelWriter(output_file) as writer:
                for name, table in tables.items():
                    table.to_excel(writer, sheet_name=name[:31], index=False)
        else:
            output_file = save_dir / 'statistical_report.tex'
            sections = []
            for name, table in tables.items():
                sections.append(f'% {name}')
                sections.append(table.to_latex(index=False, float_format='%.4g'))
            output_file.write_text('\n'.join(sections), encoding='utf-8')

        print(f"Statistical report saved to: {output_file}")
        return output_file

    def generate_convergence_plots(self) -> None:
        """
        Generate and save convergence curve plots.

        Returns
        -------
        None
            Saves figures to disk at configured save_path.

        Raises
        ------
        RuntimeError
            If calculate_metrics() has not been called.
        """
        if self._metric_results is None:
            self.calculate_metrics()

        algo_order = self.algorithm_order if self.algorithm_order else self._scan_result.algorithms

        plot_gen = PlotGenerator(self.plot_config)
        plot_gen.plot_convergence_curves(
            self._metric_results.metric_values,
            self._metric_results.best_values,
            self._metric_results.max_nfes,
            algo_order,
            self._metric_results.metric_name
        )

    def generate_runtime_plots(self) -> None:
        """
        Generate and save runtime comparison bar plots.

        Returns
        -------
        None
            Saves figure to disk at configured save_path.

        Raises
        ------
        RuntimeError
            If calculate_metrics() has not been called.
        """
        if self._metric_results is None:
            self.calculate_metrics()

        algo_order = self.algorithm_order if self.algorithm_order else self._scan_result.algorithms

        plot_gen = PlotGenerator(self.plot_config)
        plot_gen.plot_runtime(self._metric_results.runtime, algo_order)

    def generate_nd_solution_plots(self) -> None:
        """
        Generate and save non-dominated solution visualization plots.

        Returns
        -------
        None
            Saves figures to disk at configured save_path/ND_Solutions/.

        Raises
        ------
        RuntimeError
            If calculate_metrics() has not been called.
        """
        if self._metric_results is None:
            self.calculate_metrics()

        algo_order = self.algorithm_order if self.algorithm_order else self._scan_result.algorithms

        plot_gen = PlotGenerator(self.plot_config)
        plot_gen.plot_nd_solutions(
            self._metric_results.best_values,
            self._metric_results.objective_values,
            algo_order,
            self.settings
        )

    def export_convergence_data(self, k: Optional[int] = None) -> None:
        """
        Export convergence curve data to text files.

        For each problem-task combination, exports a file containing evaluation
        counts paired with convergence values for all algorithms.

        Parameters
        ----------
        k : Optional[int], optional
            Number of data points to sample from each convergence curve.
            If None, uses self.convergence_k. If both are None, exports all points.
        """
        if self._metric_results is None:
            self.calculate_metrics()

        k = k if k is not None else self.convergence_k
        algo_order = self.algorithm_order if self.algorithm_order else self._scan_result.algorithms
        metric_values = self._metric_results.metric_values
        best_values = self._metric_results.best_values
        max_nfes = self._metric_results.max_nfes
        problems = sorted(metric_values[algo_order[0]].keys())

        save_dir = Path(self.plot_config.save_path) / 'Convergence_Data'
        save_dir.mkdir(parents=True, exist_ok=True)

        plot_gen = PlotGenerator(self.plot_config)

        for prob_idx, prob in enumerate(problems):
            # Use any available run (run 1 may be missing if it failed)
            first_run_data = next(iter(best_values[algo_order[0]][prob].values()))
            num_tasks = len(first_run_data)

            for task_idx in range(num_tasks):
                filename = f'Problem{prob_idx + 1}_task{task_idx + 1}.txt'
                filepath = save_dir / filename

                with open(filepath, 'w', encoding='utf-8') as f:
                    for algo in algo_order:
                        selected_run = StatisticsCalculator.select_representative_run(
                            best_values, algo, prob, task_idx,
                            self.plot_config.statistic_type
                        )
                        curve = plot_gen._get_convergence_curve(
                            metric_values, algo, prob, task_idx, selected_run
                        )
                        if len(curve) == 0:
                            continue

                        nfes = max_nfes[algo][prob][task_idx]
                        x = np.linspace(0, nfes, len(curve))

                        # Sample k points if requested
                        if k is not None and k > 0 and len(curve) > k:
                            indices = np.linspace(0, len(curve) - 1, k, dtype=int)
                            x = x[indices]
                            curve = curve[indices]

                        f.write(f'# Algorithm: {algo}\n')
                        f.write('# NFEs\tValue\n')
                        for xi, yi in zip(x, curve):
                            f.write(f'{float(xi):.6g}\t{float(yi):.6g}\n')
                        f.write('\n')

        print(f"Convergence data exported to: {save_dir}")

    def run(self) -> MetricResults:
        """
        Execute the complete analysis pipeline.

        This method runs all analysis steps in sequence:

        1. Clear existing results (if configured)
        2. Scan data directory
        3. Calculate metrics
        4. Generate statistical tables
        5. Run the multi-problem statistical analysis (if configured)
        6. Draw the critical difference diagram (if configured)
        7. Generate convergence plots
        8. Generate runtime plots
        9. Generate non-dominated solution plots

        Returns
        -------
        MetricResults
            Complete metric results from the analysis.
        """
        print("=" * 60)
        print('🚀🚀🚀 Starting Data Analysis Pipeline! 🚀🚀🚀')
        print("=" * 60)

        # Step 0: Clear results folder if requested
        if self.clear_results:
            results_path = self.table_config.save_path
            if results_path.exists():
                print(f'\n♻️  Clearing existing results folder: {results_path}')
                shutil.rmtree(results_path)
            results_path.mkdir(parents=True, exist_ok=True)

        # Step 1: Scan data
        print('\n🔍 Scanning data directory...')
        self.scan_data()

        # Step 2: Calculate metrics
        print('\n📊 Calculating metric values...')
        self.calculate_metrics()

        # Step 3: Generate tables
        print('\n📋 Generating statistical tables...')
        self.generate_tables()

        # Step 3.5: Multi-problem statistical analysis
        if self.multi_problem_report:
            print('\n🔬 Running the multi-problem statistical analysis...')
            self.generate_statistical_report()

        if self.cd_diagram:
            print('\n📐 Drawing critical difference diagram...')
            self.generate_cd_diagram()

        # Step 4: Plot convergence curves
        print('\n📈 Plotting convergence curves...')
        self.generate_convergence_plots()

        # Step 4.5: Export convergence data
        if self.convergence_k is not None:
            print('\n📂 Exporting convergence data...')
            self.export_convergence_data()

        # Step 5: Plot runtime
        print('\n⏱️ Plotting runtime comparison...')
        self.generate_runtime_plots()

        # Step 6: Plot non-dominated solutions
        print('\n🎯 Plotting non-dominated solutions...')
        self.generate_nd_solution_plots()

        print("=" * 60)
        print('🎉🎉🎉 Data Analysis Completed! 🎉🎉🎉')
        print("=" * 60)

        return self._metric_results


# =============================================================================
# Module Entry Point and Usage Examples
# =============================================================================

if __name__ == '__main__':
    """
    Usage Examples for DataAnalyzer Module
    ======================================

    This module provides a comprehensive analysis pipeline for multi-task
    optimization experiments. Below are various usage patterns.


    Example 1: Quick Start - Full Pipeline
    --------------------------------------
    Run complete analysis with default settings::

        from data_analyzer import DataAnalyzer

        analyzer = DataAnalyzer(data_path='./Data')
        results = analyzer.run()


    Example 2: Multi-Objective Optimization with Custom Settings
    ------------------------------------------------------------
    Analyze multi-objective results with IGD metric::

        from data_analyzer import DataAnalyzer

        # Define problem settings with Pareto front references
        SETTINGS = {
            'metric': 'IGD',
            'ref_path': './MOReference',
            'n_ref': 10000,
            'P1': {
                'T1': 'P1_T1_ref.npy',
                'T2': 'P1_T2_ref.npy',
            },
            'P2': {
                'T1': lambda n, m: generate_pf(n, m),  # Callable reference
            }
        }

        analyzer = DataAnalyzer(
            data_path='./Data',
            settings=SETTINGS,
            save_path='./Results',
            table_format='latex',
            figure_format='pdf'
        )
        results = analyzer.run()


    Example 3: Step-by-Step Analysis
    --------------------------------
    Execute individual analysis steps for fine-grained control::

        from data_analyzer import DataAnalyzer

        analyzer = DataAnalyzer(
            data_path='./Data',
            settings=SETTINGS,
            algorithm_order=['NSGA-II', 'MOEA/D', 'MyAlgo'],  # Last is baseline
            clear_results=False
        )

        # Step 1: Scan data directory
        scan_result = analyzer.scan_data()
        print(f"Found algorithms: {scan_result.algorithms}")
        print(f"Found problems: {scan_result.problems}")

        # Step 2: Calculate metrics
        metric_results = analyzer.calculate_metrics()

        # Step 3: Generate only specific outputs
        analyzer.generate_tables()           # Statistical comparison tables
        analyzer.generate_convergence_plots() # Convergence curves
        analyzer.generate_runtime_plots()     # Runtime bar charts
        analyzer.generate_nd_solution_plots() # Pareto front visualizations


    Example 4: Custom Table Generation
    ----------------------------------
    Generate tables with specific statistical settings::

        from data_analyzer import (
            DataAnalyzer, TableGenerator, TableConfig,
            TableFormat, StatisticType
        )

        # Create custom table configuration
        table_config = TableConfig(
            table_format=TableFormat.LATEX,
            statistic_type=StatisticType.MEDIAN,
            significance_level=0.01,
            rank_sum_test=True,
            save_path=Path('./CustomResults')
        )

        # Use with analyzer
        analyzer = DataAnalyzer(data_path='./Data', settings=SETTINGS)
        analyzer.scan_data()
        analyzer.calculate_metrics()

        # Generate table with custom config
        table_gen = TableGenerator(table_config)
        latex_table = table_gen.generate(
            analyzer._metric_results.best_values,
            algorithm_order=['Algo1', 'Algo2', 'Baseline'],
            metric_name='IGD'
        )


    Example 5: Custom Plot Generation
    ---------------------------------
    Create plots with specific visual settings::

        from data_analyzer import DataAnalyzer, PlotGenerator, PlotConfig, StatisticType

        # Create custom plot configuration
        plot_config = PlotConfig(
            figure_format='png',
            statistic_type=StatisticType.MEDIAN,
            log_scale=True,
            show_pf=True,
            show_nd=True,
            save_path=Path('./Figures'),
            colors=['#E41A1C', '#377EB8', '#4DAF4A'],  # Custom colors
            markers=['o', 's', '^']
        )

        analyzer = DataAnalyzer(data_path='./Data', settings=SETTINGS)
        analyzer.calculate_metrics()

        # Generate plots with custom config
        plot_gen = PlotGenerator(plot_config)
        plot_gen.plot_convergence_curves(
            analyzer._metric_results.metric_values,
            analyzer._metric_results.best_values,
            analyzer._metric_results.max_nfes,
            algorithm_order=['Algo1', 'Algo2'],
            metric_name='IGD'
        )


    Example 6: Customizing Plot Font Sizes and Legend
    -------------------------------------------------
    Control font sizes, line styles, and legend appearance in convergence plots.

    The following parameters can be customized in PlotGenerator._create_convergence_figure():

    Font sizes (hardcoded, modify in source if needed)::

        ax.set_xlabel('NFEs', fontsize=14)      # X-axis label font size
        ax.set_ylabel(y_label, fontsize=14)     # Y-axis label font size
        ax.set_title(title, fontsize=14)        # Title font size
        ax.tick_params(..., labelsize=14)       # Tick label font size

    Line width and marker size (adaptive based on algorithm count)::

        # 1-4 algorithms:  markersize=8, linewidth=2.5
        # 5-6 algorithms:  markersize=7, linewidth=2.0
        # 7+ algorithms:   markersize=6, linewidth=1.6

    Legend font size (adaptive, see _calculate_legend_fontsize())::

        # 2 algorithms:   fontsize=14
        # 15 algorithms:  fontsize=6
        # Linear interpolation for values in between

    Merged plot legend (fixed size)::

        # In _plot_combined_convergence_for_problem():
        legend_fontsize = 18                    # Fixed legend font size
        legend_padding_cm = 0.3                 # Gap between legend and plots (cm)
        n_legend_cols = min(len(algorithms), 7) # Max 7 columns in legend

    To modify these values, edit the corresponding methods in PlotGenerator class.


    Example 7: Access Raw Results
    -----------------------------
    Access computed metrics for custom analysis::

        from data_analyzer import DataAnalyzer

        analyzer = DataAnalyzer(data_path='./Data', settings=SETTINGS)
        results = analyzer.run()

        # Access metric values
        # Structure: results.metric_values[algo][problem][run][task_idx]
        algo1_p1_run1_task0 = results.metric_values['Algo1']['P1'][1][0]

        # Access best values
        # Structure: results.best_values[algo][problem][run] = [task0_val, task1_val, ...]
        best_vals = results.best_values['Algo1']['P1'][1]

        # Access objective values (Pareto solutions)
        # Structure: results.objective_values[algo][problem][run][task_idx] = np.ndarray
        pareto_solutions = results.objective_values['Algo1']['P1'][1][0]

        # Access runtime
        runtime_seconds = results.runtime['Algo1']['P1'][1]

        # Access max NFEs per task
        max_nfes_list = results.max_nfes['Algo1']['P1']


    Example 8: Using Utility Classes Directly
    -----------------------------------------
    Use statistics and data utilities independently::

        from data_analyzer import (
            StatisticsCalculator, DataUtils,
            StatisticType, OptimizationDirection
        )
        import numpy as np

        # Calculate statistics
        data = [1.2, 1.5, 1.1, 1.3, 1.4]
        mean, std = StatisticsCalculator.calculate_statistic(data, StatisticType.MEAN)

        # Perform statistical comparison
        algo_data = [1.0, 1.1, 0.9, 1.2]
        base_data = [2.0, 2.1, 1.9, 2.2]
        result = StatisticsCalculator.perform_rank_sum_test(
            algo_data, base_data,
            significance_level=0.05,
            direction=OptimizationDirection.MINIMIZE
        )
        print(f"Comparison: {result.symbol}, p-value: {result.p_value}")

        # Load reference data
        reference = DataUtils.load_reference(
            settings=SETTINGS,
            problem='P1',
            task_identifier=0,  # or 'T1'
            n_objectives=2
        )


    Data Directory Structure
    ------------------------
    Expected directory structure for input data::

        ./Data/
        ├── Algorithm1/
        │   ├── Algorithm1_Problem1_1.pkl
        │   ├── Algorithm1_Problem1_2.pkl
        │   ├── Algorithm1_Problem2_1.pkl
        │   └── ...
        ├── Algorithm2/
        │   ├── Algorithm2_Problem1_1.pkl
        │   └── ...
        └── ...

    Each .pkl file should contain a dictionary with keys:

    - 'all_objs': List[List[np.ndarray]] - Objectives per task per generation
    - 'runtime': float - Total runtime in seconds
    - 'max_nfes': List[int] - Max function evaluations per task


    Output Structure
    ----------------
    Generated output files::

        ./Results/
        ├── results_table_mean.xlsx      # or .tex for LaTeX
        ├── Problem1.pdf                 # Convergence plot (single task)
        ├── Problem2-Task1.pdf           # Convergence plot (multi-task)
        ├── Problem2-Task2.pdf
        ├── runtime_comparison.pdf       # Runtime bar chart
        └── ND_Solutions/
            ├── Problem1-Algorithm1.pdf  # Pareto front plot
            ├── Problem1-Algorithm2.pdf
            └── ...
    """

    # Demo: Run analysis with sample configuration
    print("DataAnalyzer Module - Demo Run")
    print("=" * 50)

    # Example configuration (modify paths as needed)
    analyzer = DataAnalyzer(
        data_path='./Data',
        save_path='./Results',
        table_format='excel',
        figure_format='pdf',
        statistic_type='mean',
        significance_level=0.05,
        rank_sum_test=True,
        log_scale=False,
        show_pf=True,
        show_nd=True,
        clear_results=True
    )

    # Run complete analysis pipeline
    results = analyzer.run()
